import io
import json
import datetime
from datetime import datetime
#django
from django.urls import reverse
from django.db.models import Q, Sum, Min, Max
from django.http import HttpResponse
from django.db import transaction, IntegrityError
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.forms import formset_factory, inlineformset_factory
#third
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
#local
from executieves.models import Executive
from exporting.models import ExportItem
from main.decorators import role_required
from main.functions import generate_form_errors, get_auto_id, get_current_role, has_group
from profit.views import calculate_profit, profit_calculation
from purchase.models import Purchase, PurchaseExpense, PurchaseItems, PurchaseStock, PurchasedItems
from purchase.forms import PurchaseExpenseForm, PurchaseForm, PurchaseItemForm, PurchasedItemsForm
from sales.models import SalesItems

# Create your views here.
@login_required
@role_required(['superadmin','purchase','director'])
def purchase_items(request):
    """
    Purchase Items
    :param request:
    :return: Purchase Items List
    """
    instances = PurchaseItems.objects.filter(is_deleted=False)
    
    context = {
        'instances': instances,
        'page_name' : 'Purchase Items',
        'page_title' : 'Purchase Items',
        'is_purchase_pages' : True,
        'is_purchase_items_page': True,
    }

    return render(request, 'admin_panel/pages/purchase/purchase_items/list.html', context)

@login_required
@role_required(['superadmin','purchase','director'])
def create_purchase_items(request):
    PurchaseItemsFormset = formset_factory(PurchaseItemForm, extra=2)
    
    message = ''
    if request.method == 'POST':
        purchase_items_formset = PurchaseItemsFormset(request.POST,prefix='purchase_items_formset', form_kwargs={'empty_permitted': False})
        
        if purchase_items_formset.is_valid() :
            for form in purchase_items_formset:
                data = form.save(commit=False)
                data.auto_id = get_auto_id(PurchaseItems)
                data.creator = request.user
                data.save()
                    
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Product Item created successfully.",
                'redirect': 'true',
                "redirect_url": reverse('purchase:purchase_items')
            }
    
        else:
            message = generate_form_errors(purchase_items_formset, formset=True)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        purchase_items_formset = PurchaseItemsFormset(prefix='purchase_items_formset')
        
        context = {
            'purchase_items_formset': purchase_items_formset,
            
            'page_title': 'Create Purchase Items',
            'url': reverse('purchase:create_purchase_items'),
            'is_purchase_pages' : True,
            'is_purchase_items_page': True,
        }
        
        return render(request,'admin_panel/pages/purchase/purchase_items/create.html',context)
    
@login_required
@role_required(['superadmin','purchase','director'])
def edit_purchase_item(request,pk):
    """
    edit operation of purchase item
    :param request:
    :param pk:
    :return:
    """
    instance = get_object_or_404(PurchaseItems, pk=pk)
        
    message = ''
    if request.method == 'POST':
        form = PurchaseItemForm(request.POST,instance=instance)
        
        if form.is_valid():
            
            data = form.save(commit=False)
            data.date_updated = datetime.today()
            data.updater = request.user
            data.save()
                    
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Purchase Item Update successfully.",
                'redirect': 'true',
                "redirect_url": reverse('purchase:purchase_items')
            }
    
        else:
            message = generate_form_errors(form ,formset=False)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        
        form = PurchaseItemForm(instance=instance)

        context = {
            'form': form,
            'page_name' : 'Edit Purchase Items',
            'page_title' : 'Edit Purchase Items',
            'url' : reverse('purchase:create_purchase_items'),
            'is_purchase_pages' : True,
            'is_purchase_items_page': True,  
        }

        return render(request, 'admin_panel/pages/purchase/purchase_items/edit.html',context)

@login_required
@role_required(['superadmin', 'purchase', 'director'])
def delete_purchase_items(request, pk):
    """
    Purchase items deletion, it only marks the is_deleted field to true
    :param request:
    :param pk:
    :return:
    """
    export_items_exist = ExportItem.objects.filter(is_deleted=False, purchasestock__purchase_item__pk=pk).exists()
    sales_items_exist = SalesItems.objects.filter(is_deleted=False, sales_stock__purchase_item__pk=pk).exists()
    
    if not export_items_exist and not sales_items_exist:
        purchase_item = get_object_or_404(PurchaseItems, pk=pk)
        purchase_item.is_deleted = True
        purchase_item.save()
            
        response_data = {
            "statusCode": 200,
            "status": "true",
            "title": "Success",
            "message": "Purchase items deleted successfully.",
        }
    else:
        message = "This item has already been used in export or sales entries, so it can't be deleted."
        response_data = {
            "statusCode": 400,
            "status": "false",
            "title": "Failed",
            "message": message,
        }

    return HttpResponse(json.dumps(response_data), content_type='application/javascript')

@login_required
@role_required(['superadmin','purchase','director'])
def purchase(request,pk):
    """
    Purchase
    :param request:
    :return: Purchase Report single view
    """
    instance = Purchase.objects.get(pk=pk,is_deleted=False)
    items_instances = PurchasedItems.objects.filter(purchase=instance,is_deleted=False)
    expense_instances = PurchaseExpense.objects.filter(purchase=instance,is_deleted=False)
    
    context = {
        'instance': instance,
        'items_instances': items_instances,
        'expense_instances': expense_instances,
        
        'page_name' : 'Purchase Report',
        'page_title' : 'Purchase Report',
        'is_purchase_pages' : True,
        'is_purchase_page': True,
    }

    return render(request, 'admin_panel/pages/purchase/purchases/info.html', context)

@login_required
@role_required(['superadmin','purchase','investor'])
def purchase_reports(request):
    """
    Purchase Report
    :param request:
    :return: Purchase Report list view
    """
    filter_data = {}
    query = request.GET.get("q")
    
    instances = Purchase.objects.filter(is_deleted=False)
         
    date_range = request.GET.get('date_range')

    if date_range:
        start_date_str, end_date_str = date_range.split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        instances = instances.filter(date__range=[start_date, end_date])
    
    if query:

        instances = instances.filter(
            Q(invoice_no__icontains=query) |
            Q(purchase_id__icontains=query) 
        )
        title = "Purchase Report - %s" % query
        filter_data['q'] = query
        
    first_date_added = instances.aggregate(first_date_added=Min('date'))['first_date_added']
    last_date_added = instances.aggregate(last_date_added=Max('date'))['last_date_added']
    
    first_date_formatted = first_date_added.strftime('%m/%d/%Y') if first_date_added else None
    last_date_formatted = last_date_added.strftime('%m/%d/%Y') if last_date_added else None
    
    context = {
        'instances': instances.order_by("-date_added"),
        'page_name' : 'Purchase Report',
        'page_title' : 'Purchase Report',
        'filter_data' :filter_data,
        'date_range': date_range,
        'first_date_formatted': first_date_formatted,
        'last_date_formatted': last_date_formatted,
        
        'is_purchase_pages' : True,
        'is_purchase_page': True,
        'is_need_datetime_picker': True,
    }

    return render(request, 'admin_panel/pages/purchase/purchases/list.html', context)

@login_required
@role_required(['superadmin','purchase','director'])
def create_purchase(request):
    ItemsFormset = formset_factory(PurchasedItemsForm, extra=2)
    PurchaseExpensesFormset = formset_factory(PurchaseExpenseForm, extra=2)
    
    message = ''
    if request.method == 'POST':
        purchase_form = PurchaseForm(request.POST)
        purchase_items_formset = ItemsFormset(request.POST,prefix='purchase_items_formset', form_kwargs={'empty_permitted': False})
        purchase_expense_formset = PurchaseExpensesFormset(request.POST,prefix='purchase_expense_formset', form_kwargs={'empty_permitted': False})
        
        if purchase_form.is_valid() and purchase_items_formset.is_valid() and purchase_expense_formset.is_valid():
            
            # Generate a unique invoice number based on current date and a random number
            # date_part = timezone.now().strftime('%Y%m%d')
            # random_part = str(random.randint(1000, 9999))
            # invoice_number = f'IEEIP-{date_part}-{random_part}'
            executive = None
            if not request.user.is_superuser:
                executive = Executive.objects.get(user=request.user,is_deleted=False)
            
            auto_id = get_auto_id(Purchase)
            purchase_id = "IEEIP" + str(auto_id).zfill(3)    
            
            try:
                with transaction.atomic():
                    
                    if purchase_form.is_valid():
                        purchase_data = purchase_form.save(commit=False)
                        purchase_data.auto_id = auto_id
                        purchase_data.creator = request.user
                        purchase_data.executive = executive
                        purchase_data.purchase_id = purchase_id
                        purchase_data.save()
                    
                    for form in purchase_items_formset:
                        item_data = form.save(commit=False)
                        item_data.auto_id = get_auto_id(PurchasedItems)
                        item_data.creator = request.user
                        item_data.purchase = purchase_data
                        item_data.save()
                        
                        if (update_purchase_stock:=PurchaseStock.objects.filter(purchase_item=item_data.purchase_item,is_deleted=False)).exists():
                            stock = PurchaseStock.objects.get(purchase_item=item_data.purchase_item,is_deleted=False)
                            stock.qty += item_data.qty
                            if not update_purchase_stock.filter(purchase=purchase_data).exists():
                                stock.purchase.add(purchase_data)
                            stock.save()
                        else:
                            purchase_item = PurchaseItems.objects.get(pk=item_data.purchase_item.pk)
                            stock_item = PurchaseStock.objects.create(
                                auto_id = get_auto_id(PurchaseStock),
                                creator = request.user,
                                purchase_item = purchase_item,
                                qty = item_data.qty,
                            )
                            stock_item.purchase.add(purchase_data)
                            stock_item.save()
                    
                    for form in purchase_expense_formset:
                        expense_data = form.save(commit=False)
                        expense_data.auto_id = get_auto_id(PurchaseExpense)
                        expense_data.creator = request.user
                        expense_data.purchase = purchase_data
                        expense_data.save()
                        
                    # calculate_profit(purchase_data.date)
                    profit_calculation(purchase_data.date)
                        
                    response_data = {
                        "status": "true",
                        "title": "Successfully Created",
                        "message": "Product created successfully.",
                        'redirect': 'true',
                        "redirect_url": reverse('purchase:purchase_reports')
                    }
            
            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(purchase_form,formset=False)
            message += generate_form_errors(purchase_items_formset,formset=True)
            message += generate_form_errors(purchase_expense_formset,formset=True)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        purchase_form = PurchaseForm()
        purchase_items_formset = ItemsFormset(prefix='purchase_items_formset')
        purchase_expense_formset = PurchaseExpensesFormset(prefix='purchase_expense_formset')
        
        context = {
            'purchase_form': purchase_form,
            'purchase_items_formset': purchase_items_formset,
            'purchase_expense_formset': purchase_expense_formset,
            
            'page_title': 'Create Purchase',
            'url': reverse('purchase:create_purchase'),
            'is_purchase_pages' : True,
            'is_purchase_page': True,
            'is_need_datetime_picker': True,
        }
        
        return render(request,'admin_panel/pages/purchase/purchases/create.html',context)


@login_required
@role_required(['superadmin', 'purchase', 'director'])
def edit_purchase(request, pk):
    """
    Edit operation of purchase
    :param request:
    :param pk:
    :return:
    """
    purchase_instance = get_object_or_404(Purchase, pk=pk)
    purchased_items = PurchasedItems.objects.filter(purchase=purchase_instance)
    expenses = PurchaseExpense.objects.filter(purchase=purchase_instance)

    if PurchasedItems.objects.filter(purchase=purchase_instance).exists():
        i_extra = 0
    else:
        i_extra = 1

    if PurchaseExpense.objects.filter(purchase=purchase_instance).exists():
        e_extra = 0
    else:
        e_extra = 1

    ItemsFormset = inlineformset_factory(
        Purchase,
        PurchasedItems,
        extra=i_extra,
        form=PurchasedItemsForm,
    )

    ExpensesFormset = inlineformset_factory(
        Purchase,
        PurchaseExpense,
        extra=e_extra,
        form=PurchaseExpenseForm,
    )

    message = ''

    if request.method == 'POST':
        for purchased_item in purchased_items:
            if (update_purchase_stock:=PurchaseStock.objects.filter(purchase_item=purchased_item.purchase_item,is_deleted=False)).exists():
                stock = PurchaseStock.objects.get(purchase_item=purchased_item.purchase_item,is_deleted=False)
                stock.qty -= purchased_item.qty
                if not update_purchase_stock.filter(purchase=purchase_instance).exists():
                    stock.purchase.add(purchase_instance)
                stock.save()
            
        purchased_items.update(
            qty = 0,
            amount = 0,
            amount_per_kg = 0
        )
        expenses.update(
            amount = 0,
        )
        
        purchase_form = PurchaseForm(request.POST, instance=purchase_instance)
        purchase_items_formset = ItemsFormset(request.POST, request.FILES,
                                              instance=purchase_instance,
                                              prefix='purchase_items_formset',
                                              form_kwargs={'empty_permitted': False})
        purchase_expense_formset = ExpensesFormset(request.POST,
                                                   instance=purchase_instance,
                                                   prefix='purchase_expense_formset',
                                                   form_kwargs={'empty_permitted': False})

        if purchase_form.is_valid() and purchase_items_formset.is_valid() and purchase_expense_formset.is_valid():
            try:
                with transaction.atomic():
                    # Update purchase data
                    purchase_instance = purchase_form.save(commit=False)
                    purchase_instance.date_updated = datetime.today().now()
                    purchase_instance.updater = request.user
                    purchase_instance.date = request.POST.get('date')
                    purchase_instance.save()

                    # Update purchase items
                    for form in purchase_items_formset:
                        if form not in purchase_items_formset.deleted_forms:
                            item_data = form.save(commit=False)
                            if not item_data.auto_id:
                                item_data.purchase = purchase_instance
                                item_data.auto_id = get_auto_id(PurchasedItems)
                                item_data.creator = request.user
                                
                                if (update_purchase_stock:=PurchaseStock.objects.filter(purchase_item=item_data.purchase_item,is_deleted=False)).exists():
                                    stock = PurchaseStock.objects.get(purchase_item=item_data.purchase_item,is_deleted=False)
                                    stock.qty += item_data.qty
                                    if not update_purchase_stock.filter(purchase=purchase_instance).exists():
                                        stock.purchase.add(purchase_instance)
                                    stock.save()
                                else:
                                    purchase_item = PurchaseItems.objects.get(pk=item_data.purchase_item.pk)
                                    stock_item = PurchaseStock.objects.create(
                                        auto_id = get_auto_id(PurchaseStock),
                                        creator = request.user,
                                        purchase_item = purchase_item,
                                        qty = item_data.qty,
                                    )
                                    stock_item.purchase.add(purchase_instance)
                                    stock_item.save()
                            item_data.save()

                            # Adjust stock quantities
                            update_stock_quantity(request,item_data)

                    # Delete removed purchase items
                    for form in purchase_items_formset.deleted_forms:
                        form.instance.delete()

                    # Update purchase expenses
                    for form in purchase_expense_formset:
                        if form not in purchase_expense_formset.deleted_forms:
                            expense_data = form.save(commit=False)
                            if not expense_data.auto_id:
                                expense_data.auto_id = get_auto_id(PurchaseExpense)
                                expense_data.creator = request.user
                                expense_data.purchase = purchase_instance
                            expense_data.save()

                    # Delete removed purchase expenses
                    for form in purchase_expense_formset.deleted_forms:
                        form.instance.delete()

                    # Calculate profit
                    # calculate_profit(purchase_instance.date)
                    profit_calculation(datetime.strptime(purchase_instance.date, '%Y-%m-%d').date())

                    response_data = {
                        "status": "true",
                        "title": "Successfully Updated",
                        "message": "Purchase updated successfully.",
                        "redirect": "true",
                        "redirect_url": reverse('purchase:purchase_reports'),
                        "return": True,
                    }

            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

        else:
            message = generate_form_errors(purchase_form, formset=False)
            message += generate_form_errors(purchase_items_formset, formset=True)
            message += generate_form_errors(purchase_expense_formset, formset=True)

            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')

    else:
        purchase_form = PurchaseForm(instance=purchase_instance)
        purchase_items_formset = ItemsFormset(queryset=purchased_items,
                                               prefix='purchase_items_formset',
                                               instance=purchase_instance)
        purchase_expense_formset = ExpensesFormset(queryset=expenses,
                                                   prefix='purchase_expense_formset',
                                                   instance=purchase_instance)

        context = {
            'purchase_form': purchase_form,
            'purchase_items_formset': purchase_items_formset,
            'purchase_expense_formset': purchase_expense_formset,

            'message': message,
            'page_name': 'edit purchase',
            'is_purchase_pages': True,
            'is_purchase_page': True,
        }

        return render(request, 'admin_panel/pages/purchase/purchases/create.html', context)


def update_stock_quantity(request, purchased_item):
    """
    Update stock quantity based on purchased item.
    """
    if (update_purchase_stock := PurchaseStock.objects.filter(purchase_item=purchased_item.purchase_item,
                                                               is_deleted=False)).exists():
        stock = update_purchase_stock.first()
        stock.qty += purchased_item.qty
        stock.save()
    else:
        purchase_item = PurchaseItems.objects.get(pk=purchased_item.purchase_item.pk)
        PurchaseStock.objects.create(
            auto_id=get_auto_id(PurchaseStock),
            creator=request.user,
            purchase_item=purchase_item,
            qty=purchased_item.qty,
        )

    
@login_required
@role_required(['superadmin', 'purchase', 'director'])
def delete_purchase(request, pk):
    try:
        with transaction.atomic():
            purchase_instance = Purchase.objects.get(pk=pk)
            purchase_items = PurchasedItems.objects.filter(purchase=purchase_instance,is_deleted=False)
            purchase_expenses = PurchaseExpense.objects.filter(purchase=purchase_instance,is_deleted=False)
            
            for item in purchase_items:
                if (stock_instance:=PurchaseStock.objects.filter(purchase_item=item.purchase_item,qty__gte=item.qty)).exists():
                    stock_instance = stock_instance.first()
                    stock_instance.qty -= item.qty
                    stock_instance.save()
                
                    item.is_deleted = True
                    item.save()
                else:
                    response_data = {
                        "StatusCode": 400,
                        "status": "false",
                        "title": "Failed",
                        "message": f"Item {item.purchase_item.name}",
                    }
            
            purchase_expenses.update(is_deleted=True)
            
            purchase_instance.is_deleted=True
            purchase_instance.save()
            
            # calculate_profit(purchase_instance.date)
            profit_calculation(purchase_instance.date)
            
            response_data = {
                "StatusCode": 200,
                "status": "true",
                "title": "Success",
                "message": "Successfully Deleted",
            }
            
    except IntegrityError as e:
        response_data = {
            "StatusCode": 400,
            "status": "false",
            "title": "Failed",
            "message": str(e),
        }

    except Exception as e:
        response_data = {
            "StatusCode": 400,
            "status": "false",
            "title": "Failed",
            "message": str(e),
        }

    return HttpResponse(json.dumps(response_data), content_type='application/javascript')

@login_required
@role_required(['superadmin','purchase','director'])
def edit_purchased_item(request,pk):
    """
    edit operation of purchased item
    :param request:
    :param pk:
    :return:
    """
    instance = get_object_or_404(PurchasedItems, pk=pk)
        
    message = ''
    if request.method == 'POST':
        instance.amount = 0
        instance.save()
        
        form = PurchasedItemsForm(request.POST,instance=instance)
        
        if form.is_valid():
            
            data = form.save(commit=False)
            data.date_updated = datetime.today()
            data.updater = request.user
            data.save()
            
            # Adjust stock quantities
            update_stock_quantity(data)
            # Calculate profit
            # calculate_profit(data.purchase.date)
            profit_calculation(data.purchase.date)
                    
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Purchased Item Update successfully.",
                'redirect': 'true',
                "redirect_url": reverse('purchase:purchase_reports')
            }
    
        else:
            message = generate_form_errors(form ,formset=False)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        
        form = PurchasedItemsForm(instance=instance)

        context = {
            'form': form,
            'page_name' : 'Edit Purchased Items',
            'page_title' : 'Edit Purchased Items',
            'is_purchase_pages' : True,
            'is_purchase_items_page': True,  
        }

        return render(request, 'admin_panel/pages/create/create.html',context)

@login_required
@role_required(['superadmin','purchase','director'])
def print_purchases(request):
    """
    Purchase Print
    :param request:
    :return: Purchase Print List
    """
    purchase_list = Purchase.objects.filter(is_deleted=False)
    
    context = {
        'purchase_list': purchase_list,
        'page_name' : 'Print Purchases',
        'page_title' : 'Print Purchases',
        'is_purchase_pages' : True,
        'is_purchase_items_page': True,
    }

    return render(request, 'admin_panel/pages/purchase/purchases/print.html', context)

def export_purchases(request):
    filter_data = {}
    purchase_pk = request.GET.get("purchase_pk")

    purchases = Purchase.objects.filter()

    if purchase_pk:
        purchases = purchases.filter(pk=purchase_pk)

    date_range = request.GET.get('date_range')
    if date_range:
        start_date_str, end_date_str = date_range.split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        purchases = purchases.filter(date__range=[start_date, end_date])
        filter_data['date_range'] = date_range
    
    query = request.GET.get("q")
    if query:
        purchases = purchases.filter(
            Q(purchase_id__icontains=query) 
        )
        filter_data['q'] = query

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column headers
    ws.append(['Purchase ID', 'Date', 'Purchase Party', 'Executive', 'Total Quantity', 'Materials Total Amount',
               'Materials Total Expense', 'Sub Total'])

    # Set column widths
    column_widths = [15, 15, 20, 20, 15, 20, 20, 15]

    for i, column_width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = column_width

    # Iterate through Purchase objects and fetch associated PurchasedItems and PurchaseExpense
    for purchase in purchases:
        purchase_party = purchase.purchase_party.get_fullname() if purchase.purchase_party else ''
        executive = purchase.executive.get_fullname() if purchase.executive else ''

        ws.append([
            purchase.purchase_id,
            purchase.date,
            purchase_party,
            executive,
            purchase.total_qty(),
            purchase.materials_total_amount(),
            purchase.materials_total_expence(),
            purchase.sub_total(),
        ])

    output = io.BytesIO()
    wb.save(output)

    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=purchases_data.xlsx'

    return response


@login_required
@role_required(['superadmin','purchase','director'])
def purchase_stock(request):
    """
    Purchase Stocks
    :param request:
    :return: Purchase Stocks List
    """
    instances = PurchaseStock.objects.filter(is_deleted=False)
    
    context = {
        'instances': instances,
        'page_name' : 'Purchase Stocks',
        'page_title' : 'Purchase Stocks',
        'is_purchase_pages' : True,
        'is_purchase_stock_page': True,
    }

    return render(request, 'admin_panel/pages/purchase/purchase_stock/list.html', context)

@login_required
@role_required(['superadmin','purchase','director'])
def print_purchase_stock(request):
    """
    Purchase stock Print
    :param request:
    :return: Purchase stock Print List
    """
    instances = PurchaseStock.objects.filter(is_deleted=False)
    
    total = 0 
    for i in instances:
        total += i.qty
    
    context = {
        'instances': instances,
        'grand_total': total,
        'page_name' : 'Print Purchase Stock',
        'page_title' : 'Purchase Stock',
        'is_purchase_pages' : True,
        'is_purchase_stock_page': True,
    }

    return render(request, 'admin_panel/pages/purchase/purchase_stock/print.html', context)

@login_required
@role_required(['superadmin','purchase','director','investor'])
def export_purchase_stock(request):
    filter_data = {}
    instances = PurchaseStock.objects.filter(is_deleted=False)

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column headers
    ws.append(['Items','Quantity'])

    # Fetch and write data to Excel
    for instance in instances:
        ws.append([
            instance.purchase_item.name,
            instance.qty
        ])

    # Prepare response
    output = io.BytesIO()
    wb.save(output)

    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=product_stock_data.xlsx'

    return response