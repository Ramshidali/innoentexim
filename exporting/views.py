import io
import json
import datetime
from datetime import datetime
#django
from django.urls import reverse
from django.http import HttpResponse
from django.db.models import Q, Sum, Min, Max
from django.db import transaction, IntegrityError
from django.contrib.auth.models import User,Group
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.forms import formset_factory, inlineformset_factory
# rest framework
from rest_framework import status
# third
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# local
from sales.models import SalesStock
from executieves.models import Executive
from purchase.models import PurchaseStock
from main.decorators import role_required
from main.functions import generate_form_errors, get_auto_id, get_current_role, has_group
from . models import CourierPartner, ExportItem, ExportStatus, Exporting, ExportingCountry
from exporting.forms import CourierPartnerForm, ExportingCountryForm, ExportingForm, ExportingItemsForm, ExportingStatusForm

# Create your views here.
def product_item_qty(request):
    item_pk = request.GET.get("purchase_item")
    
    if (instances:=PurchaseStock.objects.filter(pk=item_pk,is_deleted=False)).exists():
        qty = instances.first().qty
        
        status_code = status.HTTP_200_OK
        response_data = {
            "status": "true",
            "qty": str(qty),
        }
    else:
        status_code = status.HTTP_404_NOT_FOUND
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": "item not found",
        }

    return HttpResponse(json.dumps(response_data),status=status_code, content_type="application/json")

@login_required
@role_required(['superadmin','core_team','director'])
def export_countries(request):
    """
    Export Countries
    :param request:
    :return: Export Countries List
    """
    instances = ExportingCountry.objects.filter(is_deleted=False)
    
    context = {
        'instances': instances,
        'page_name' : 'Export Countries',
        'page_title' : 'Export Countries',
        'is_exporting' : True,
        'is_export_countries_page': True,
    }

    return render(request, 'admin_panel/pages/export/countries/list.html', context)

@login_required
@role_required(['superadmin','core_team','director'])
def create_export_country(request):
    ExportingCountryFormset = formset_factory(ExportingCountryForm, extra=2)
    
    message = ''
    if request.method == 'POST':
        export_countries_formset = ExportingCountryFormset(request.POST,prefix='export_countries_formset', form_kwargs={'empty_permitted': False})
        
        if export_countries_formset.is_valid() :
            for form in export_countries_formset:
                data = form.save(commit=False)
                data.auto_id = get_auto_id(ExportingCountry)
                data.creator = request.user
                data.save()
                    
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Product Item created successfully.",
                'redirect': 'true',
                "redirect_url": reverse('exporting:export_countries')
            }
    
        else:
            message = generate_form_errors(export_countries_formset, formset=True)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        export_countries_formset = ExportingCountryFormset(prefix='export_countries_formset')
        
        context = {
            'export_countries_formset': export_countries_formset,
            
            'page_title': 'Create Export Countries',
            'url': reverse('exporting:create_export_country'),
            'is_exporting' : True,
            'is_export_countries_page': True,
        }
        
        return render(request,'admin_panel/pages/export/countries/create.html',context)
    
@login_required
@role_required(['superadmin','core_team','director'])
def edit_export_country(request,pk):
    """
    edit operation of export item
    :param request:
    :param pk:
    :return:
    """
    instance = get_object_or_404(ExportingCountry, pk=pk)
        
    message = ''
    if request.method == 'POST':
        form = ExportingCountryForm(request.POST,instance=instance)
        
        if form.is_valid():
            
            data = form.save(commit=False)
            data.date_updated = datetime.today()
            data.updater = request.user
            data.save()
                    
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Purchase Item Update successfully.",
                'redirect': 'true',
                "redirect_url": reverse('exporting:export_countries')
            }
    
        else:
            message = generate_form_errors(form ,formset=False)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        
        form = ExportingCountryForm(instance=instance)

        context = {
            'form': form,
            'page_name' : 'Edit Export Countries',
            'page_title' : 'Edit Export Countries',
            'url' : reverse('exporting:create_export_country'),
            'is_exporting' : True,
            'is_export_countries_page': True,  
        }

        return render(request, 'admin_panel/pages/export/countries/edit.html',context)

@login_required
@role_required(['superadmin','core_team','director'])
def delete_export_country(request, pk):
    """
    export items deletion, it only mark as is deleted field to true
    :param request:
    :param pk:
    :return:
    """
    invoice_numbers_text = ""
    # data = ExportingCountry.objects.get(pk=pk)
    # if not SalesMaterials.objects.filter(export_item=data, is_deleted=False).exists():
        
    #     data.is_deleted=True
    #     data.save()
    
    #     if (instances:=PurchaseMaterials.objects.filter(export_item=data)).exists():
    #         instances.update(is_deleted=True)
            
    #     if (stocks:=Stock.objects.filter(export_item=data)).exists():
    #         stocks.update(is_deleted=True)

    #     response_data = {
    #         "status": "true",
    #         "title": "Successfully Deleted",
    #         "message": "Purchase Item Successfully Deleted.",
    #         "redirect": "true",
    #         "redirect_url": reverse('exporting:export_countries'),
    #     }
    # else:
    #     sales_material_queryset = SalesMaterials.objects.filter(export_item=data)
    #     sales_invoice_numbers = [sales_material.sales.invoice_no for sales_material in sales_material_queryset]
    #     unique_invoice_numbers = list(set(sales_invoice_numbers))
    #     invoice_numbers_text = ", ".join(unique_invoice_numbers)

    message = f"This export item has already included the sale of some items. Sales Invoice Numbers: {invoice_numbers_text}"
    
    response_data = {
        "status": "false",
        "title": "Failed",
        "message": message,
    }

    return HttpResponse(json.dumps(response_data), content_type='application/javascript')

# Create your views here.
@login_required
@role_required(['superadmin','core_team'])
def courier_partner_info(request,pk):
    """
    Courier Partner info
    :param request:
    :return: Courier Partner info view
    """
    instance = CourierPartner.objects.get(pk=pk,is_deleted=False)

    context = {
        'instance': instance,
        'page_name' : 'Courier Partner',
        'page_title' : 'Courier Partner',
        'is_exporting' : True,
        'is_courier_partner': True,
    }

    return render(request, 'admin_panel/pages/courier_partner/info.html', context)

@login_required
@role_required(['superadmin','core_team'])
def courier_partner_list(request):
    """
    Courier Partner listings
    :param request:
    :return: Courier Partner list view
    """
    instances = CourierPartner.objects.filter(is_deleted=False).order_by("-date_added")
    
    filter_data = {}
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(customer_id__icontains=query)
        )
        title = "CourierPartner - %s" % query
        filter_data['q'] = query
    

    context = {
        'instances': instances,
        'page_name' : 'Courier Partner',
        'page_title' : 'Courier Partner',
        'filter_data' :filter_data,
        'is_exporting' : True,
        'is_courier_partner': True,
    }

    return render(request, 'admin_panel/pages/export/courier_partner/list.html', context)

@login_required
@role_required(['superadmin','core_team'])
def create_courier_partner(request):
    """
    create operation of Courier Partner
    :param request:
    :param pk:
    :return:
    """
    if request.method == 'POST':
        form = CourierPartnerForm(request.POST,files=request.FILES)
            
        if form.is_valid():
            
            email = form.cleaned_data['email']
            try:
                with transaction.atomic():
                
                    user_data = User.objects.create_user(
                        username=email,
                    )
                    
                    if Group.objects.filter(name="courier_partner").exists():
                        group = Group.objects.get(name="courier_partner")
                    else:
                        group = Group.objects.create(name="courier_partner")

                    user_data.groups.add(group)
                    
                    auto_id = get_auto_id(CourierPartner)
                    regid = "IEEICP" + str(auto_id).zfill(3)
                    
                    data = form.save(commit=False)
                    data.auto_id = auto_id
                    data.creator = request.user
                    data.date_updated = datetime.today()
                    data.updater = request.user
                    data.employee_id = regid
                    data.save()
                    
                    response_data = {
                    "status": "true",
                    "title": "Successfully Created",
                    "message": "Courier Partner created successfully.",
                    'redirect': 'true',
                    "redirect_url": reverse('exporting:courier_partner_list')
                    }
                
            except IntegrityError as e:
                    
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
                
        else:
            message =generate_form_errors(form , formset=False)
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        
        form = CourierPartnerForm()

        context = {
            'form': form,
            'page_name' : 'Create Courier Partner',
            'page_title' : 'Create Courier Partner',
            'url' : reverse('exporting:create_courier_partner'),
            
            'is_need_datetime_picker': True,
            'is_need_forms': True,
            'is_exporting' : True,
            'is_courier_partner': True,
        }

        return render(request, 'admin_panel/pages/export/courier_partner/create.html',context)
    
@login_required
@role_required(['superadmin','core_team'])
def edit_courier_partner(request,pk):
    """
    edit operation of courier_partner
    :param request:
    :param pk:
    :return:
    """
    instance = get_object_or_404(CourierPartner, pk=pk)
        
    message = ''
    if request.method == 'POST':
        form = CourierPartnerForm(request.POST,files=request.FILES,instance=instance)
        
        if form.is_valid():
            #update CourierPartner
            data = form.save(commit=False)
            data.date_updated = datetime.datetime.today()
            data.updater = request.user
            data.save()
                    
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Courier Partner Update successfully.",
                'redirect': 'true',
                "redirect_url": reverse('exporting:courier_partner_list')
            }
    
        else:
            message = generate_form_errors(form ,formset=False)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        
        form = CourierPartnerForm(instance=instance)

        context = {
            'form': form,
            'page_name' : 'Edit Courier Partner',
            'page_title' : 'Edit Courier Partner',
            'is_need_datetime_picker': True,
            'is_need_forms': True,
            'is_exporting' : True,
            'is_courier_partner': True,
        }

        return render(request, 'admin_panel/pages/export/courier_partner/create.html',context)

@login_required
@role_required(['superadmin','core_team'])
def delete_courier_partner(request, pk):
    """
    CourierPartner deletion, it only mark as is deleted field to true
    :param request:
    :param pk:
    :return:
    """
    # instance = CourierPartner.objects.get(pk=pk)
    # current_email = instance.email
    # append_email = current_email + str(randomnumber(3)) + "_deleted"
    
    # instance.email = append_email
    # instance.phone = randomnumber(5)
    # instance.is_deleted = True
    # instance.save()
    
    # user = User.objects.get(username=current_email)
    # user.username = append_email
    # user.save()

    response_data = {
        "status": "true",
        "title": "Successfully Deleted",
        "message": "Courier Partner Successfully Deleted.",
        "redirect": "true",
        "redirect_url": reverse('exporting:courier_partner_list'),
    }

    return HttpResponse(json.dumps(response_data), content_type='application/javascript')

@login_required
@role_required(['superadmin','core_team','executive','director'])
def exporting(request,pk):
    """
    Exporting
    :param request:
    :return: Exporting Report single view
    """
    instance = Exporting.objects.get(pk=pk,is_deleted=False)
    items_instances = ExportItem.objects.filter(export=instance,is_deleted=False)
    
    context = {
        'instance': instance,
        'items_instances': items_instances,
        'page_name' : 'Exporting Report',
        'page_title' : 'Exporting Report',
        'is_exporting' : True,
        'is_exporting_page': True,
    }

    return render(request, 'admin_panel/pages/export/exporting/info.html', context)

# @login_required
# @role_required(['superadmin','core_team','director'])
# def print_exporting(request,pk):
#     """
#     Print Exporting
#     :param request:
#     :return: Exporting Print
#     """
#     instance = Exporting.objects.get(pk=pk,is_deleted=False)
    
#     context = {
#         'instance': instance,
#         'page_name' : 'Exporting Report',
#         'page_title' : 'Exporting Report',
#     }

#     return render(request, 'admin_panel/pages/exporting/print_exporting.html', context)

@login_required
@role_required(['superadmin','core_team','executive'])
def exporting_list(request):
    """
    Exporting Report
    :param request:
    :return: Exporting Report list view
    """
    
    filter_data = {}
    instances = Exporting.objects.filter(is_deleted=False).order_by("-date_added")
         
    date_range = ""
    date_range = request.GET.get('date_range')
    # print(date_range)

    if date_range:
        start_date_str, end_date_str = date_range.split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        instances = instances.filter(date_added__range=[start_date, end_date])
        filter_data['date_range'] = date_range
    
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(exporting_id__icontains=query) 
            # Q(exporting_id__icontains=query) 
        )
        title = "Exporting Report - %s" % query
        filter_data['q'] = query
        
    first_date_added = instances.aggregate(first_date_added=Min('date'))['first_date_added']
    last_date_added = instances.aggregate(last_date_added=Max('date'))['last_date_added']
    
    first_date_formatted = first_date_added.strftime('%m/%d/%Y') if first_date_added else None
    last_date_formatted = last_date_added.strftime('%m/%d/%Y') if last_date_added else None
    
    if request.GET.get("country_name"):
        instances = instances.filter(exporting_country__pk=request.GET.get("country_name"))
        filter_data['country_name'] = request.GET.get("country_name")
        
    if request.GET.get("courier_partner"):
        instances = instances.filter(courier_partner__pk=request.GET.get("courier_partner"))
        filter_data['courier_partner'] = request.GET.get("courier_partner")
        
    if request.GET.get("courier_status"):
        courier_status_exports = ExportStatus.objects.filter(status=request.GET.get("courier_status")).values_list("export__pk")
        instances = instances.filter(pk__in=courier_status_exports)
        filter_data['courier_status'] = request.GET.get("courier_status")
    
    context = {
        'instances': instances,
        'status_form': ExportingStatusForm(),
        'page_name' : 'Exporting Report',
        'page_title' : 'Exporting Report',
        'filter_data' :filter_data,
        'first_date_formatted': first_date_formatted,
        'last_date_formatted': last_date_formatted,
        
        'is_exporting' : True,
        'is_exporting_page': True,
        'is_need_datetime_picker': True,
    }

    return render(request, 'admin_panel/pages/export/exporting/list.html', context)

@login_required
@role_required(['superadmin','core_team','executive'])
def create_exporting(request):
    ExportingItemsFormset = formset_factory(ExportingItemsForm, extra=2)
    
    message = ''
    if request.method == 'POST':
        exporting_form = ExportingForm(request.POST)
        exporting_items_formset = ExportingItemsFormset(request.POST,prefix='exporting_items_formset', form_kwargs={'empty_permitted': False})
        
        if exporting_form.is_valid() and exporting_items_formset.is_valid() :
            executive = None
            if not request.user.is_superuser:
                executive = Executive.objects.get(user=request.user,is_deleted=False)
            
            auto_id = get_auto_id(Exporting)
            exporting_id = "IEEEXPI" + str(auto_id).zfill(3)    
            
            try:
                with transaction.atomic():
                    
                    if exporting_form.is_valid():
                        exporting_data = exporting_form.save(commit=False)
                        exporting_data.auto_id = get_auto_id(Exporting)
                        exporting_data.creator = request.user
                        exporting_data.executive = executive
                        exporting_data.exporting_id = exporting_id
                        exporting_data.save()
                    
                    for form in exporting_items_formset:
                        item_data = form.save(commit=False)
                        item_data.auto_id = get_auto_id(ExportItem)
                        item_data.creator = request.user
                        item_data.export = exporting_data
                        item_data.save()
                        
                        if (purchase_stock:=PurchaseStock.objects.filter(purchase_item=item_data.purchasestock.purchase_item)).exists():
                            purchase_stock = purchase_stock.first()
                            purchase_stock.qty -= item_data.qty
                            purchase_stock.save()
                            
                    ExportStatus.objects.create(
                        export = exporting_data,
                        creator = request.user,
                    )
                    
                    response_data = {
                        "status": "true",
                        "title": "Successfully Created",
                        "message": "Export created successfully.",
                        'redirect': 'true',
                        "redirect_url": reverse('exporting:exporting_list')
                    }
                    
            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(exporting_form,formset=False)
            message += generate_form_errors(exporting_items_formset,formset=True)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        exporting_form = ExportingForm()
        exporting_items_formset = ExportingItemsFormset(prefix='exporting_items_formset')
        
        context = {
            'exporting_form': exporting_form,
            'exporting_items_formset': exporting_items_formset,
            
            'page_title': 'Create Exporting',
            'url': reverse('exporting:create_exporting'),
            'is_exporting' : True,
            'is_exporting_page': True,
            'is_need_datetime_picker': True,
        }
        
        return render(request,'admin_panel/pages/export/exporting/create.html',context)

@login_required
@role_required(['superadmin','core_team','director','executive'])
def edit_exporting(request,pk):
    """
    edit operation of exporting
    :param request:
    :param pk:
    :return:
    """
    export_instance = get_object_or_404(Exporting, pk=pk)
    exporting_items = ExportItem.objects.filter(export=export_instance)
    
    if ExportItem.objects.filter(export=export_instance).exists():
        i_extra = 0
    else:
        i_extra = 1 
        
    ExportingItemsFormset = inlineformset_factory(
        Exporting,
        ExportItem,
        extra=i_extra,
        form=ExportingItemsForm,
    )
    
    message = ''
    
    if request.method == 'POST':
        
        if ExportStatus.objects.filter(export=export_instance,status="025").exists():
            for item in exporting_items:
                if not SalesStock.objects.get(country=item.export.exporting_country,purchase_item=item.purchasestock.purchase_item,is_deleted=False).qty > 0:
                    
                    response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": f'no stock available in {item.purchasestock.purchase_item.name} in sales stock'
                    }

                    return HttpResponse(json.dumps(response_data), content_type='application/javascript')
                
        
        for exporting_item in exporting_items:
            
            if PurchaseStock.objects.filter(purchase_item=exporting_item.purchasestock.purchase_item,is_deleted=False).exists():
                stock = PurchaseStock.objects.get(purchase_item=exporting_item.purchasestock.purchase_item,is_deleted=False)
                stock.qty += exporting_item.qty
                stock.save()
            if ExportStatus.objects.filter(export=export_instance,status="025").exists():
                if SalesStock.objects.filter(country=exporting_item.export.exporting_country,purchase_item=exporting_item.purchasestock.purchase_item,is_deleted=False).exists():
                    stock = SalesStock.objects.get(country=exporting_item.export.exporting_country,purchase_item=exporting_item.purchasestock.purchase_item,is_deleted=False)
                    stock.qty -= exporting_item.qty
                    stock.save()
        
        exporting_items.update(
            qty = 0,
        )
        
        exporting_form = ExportingForm(request.POST,instance=export_instance)
        exporting_items_formset = ExportingItemsFormset(request.POST,request.FILES,
                                            instance=export_instance,
                                            prefix='exporting_items_formset',
                                            form_kwargs={'empty_permitted': False})            
        
        if exporting_form.is_valid() and  exporting_items_formset.is_valid() :
            #create
            data = exporting_form.save(commit=False)
            data.date_updated = datetime.today()
            data.updater = request.user
            data.date = request.POST.get('date')
            data.save()
            
            for form in exporting_items_formset:
                if form not in exporting_items_formset.deleted_forms:
                    i_data = form.save(commit=False)
                    if not i_data.auto_id :
                        i_data.exporting = export_instance
                        i_data.auto_id = get_auto_id(ExportItem)
                        i_data.creator = request.user
                    i_data.save()
                    
                if (purchase_stock:=PurchaseStock.objects.filter(purchase_item=i_data.purchasestock.purchase_item)).exists():
                    purchase_stock = purchase_stock.first()
                    purchase_stock.qty -= i_data.qty
                    purchase_stock.save()

            for f in exporting_items_formset.deleted_forms:
                f.instance.delete()
                
            ExportStatus.objects.create(
                status = "010",
                caption = "export edited",
                export = export_instance,
                creator = request.user,
            )
                
            response_data = {
                "status": "true",
                "title": "Successfully Updated",
                "message": "Exporting Updated Successfully.",
                'redirect': 'true',
                "redirect_url": reverse('exporting:exporting_list'),
                "return" : True,
            }
    
        else:
            message = generate_form_errors(exporting_form,formset=False)
            message += generate_form_errors(exporting_items_formset,formset=True)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
                        
    else:
        exporting_form = ExportingForm(instance=export_instance)
        exporting_items_formset = ExportingItemsFormset(queryset=exporting_items,
                                            prefix='exporting_items_formset',
                                            instance=export_instance)

        context = {
            'exporting_form': exporting_form,
            'exporting_items_formset': exporting_items_formset,
            
            'message': message,
            'page_name' : 'edit exporting',
            'is_exporting' : True,
            'is_exporting_page': True,        
        }

        return render(request, 'admin_panel/pages/export/exporting/create.html', context)
    
@login_required
@role_required(['superadmin','core_team','director'])
def delete_exporting(request, pk):
    """
    exporting deletion, it only mark as is deleted field to true
    :param request:
    :param pk:
    :return:
    """
    
    export_instance = get_object_or_404(Exporting, pk=pk)
    exporting_items = ExportItem.objects.filter(export=export_instance)
    
    if ExportStatus.objects.filter(export=export_instance,status="025").exists():
        for item in exporting_items:
            if not SalesStock.objects.get(country=item.export.exporting_country,purchase_item=item.purchasestock.purchase_item,is_deleted=False).qty > 0:
                response_data = {
                "status": "false",
                "title": "Failed",
                "message": f'no stock available in {item.purchasestock.purchase_item.name} in sales stock'
                }

                return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    for exporting_item in exporting_items:
            
        if PurchaseStock.objects.filter(purchase_item=exporting_item.purchasestock.purchase_item,is_deleted=False).exists():
            stock = PurchaseStock.objects.get(purchase_item=exporting_item.purchasestock.purchase_item,is_deleted=False)
            stock.qty += exporting_item.qty
            stock.save()
            
        if ExportStatus.objects.filter(export=export_instance,status="025").exists():
            if SalesStock.objects.filter(country=exporting_item.export.exporting_country,purchase_item=exporting_item.purchasestock.purchase_item,is_deleted=False).exists():
                stock = SalesStock.objects.get(country=exporting_item.export.exporting_country,purchase_item=exporting_item.purchasestock.purchase_item,is_deleted=False)
                stock.qty -= exporting_item.qty
                stock.save()
            
    exporting_items.update(is_deleted=True)
    
    export_instance.is_deleted = True
    export_instance.save()
    
    response_data = {
        "status": "true",
        "title": "Successfully Deleted",
        "message": "Exporting Successfully Deleted.",
        "redirect": "true",
        "redirect_url": reverse('exporting:exporting_list'),
    }
    
    return HttpResponse(json.dumps(response_data), content_type='application/javascript')

@login_required
@role_required(['superadmin','core_team','executive'])
def update_expoting_status(request):
    """
    updation operation of exporting status
    :param request:
    :param pk:
    :return:
    """
    # print('enter')
    if request.method == 'POST':
        # print('post')
        form = ExportingStatusForm(request.POST)
        if form.is_valid():
            # print('valid')
            export = Exporting.objects.get(pk=form.cleaned_data['export_id'])
            
            data = form.save(commit=False)
            data.creator = request.user
            data.export = export
            data.save()
            
            if form.cleaned_data['status'] == "025" :
                export_items = ExportItem.objects.filter(export=export)
                for item in export_items:
                    if (update_sales_stock:=SalesStock.objects.filter(purchase_item=item.purchasestock.purchase_item,country=export.exporting_country,is_deleted=False)).exists():
                        stock = SalesStock.objects.get(purchase_item=item.purchasestock.purchase_item,country=export.exporting_country,is_deleted=False)
                        stock.qty += item.qty
                        if not update_sales_stock.filter(export=export).exists():
                            stock.export.add(export)
                        stock.save()
                    else:
                        stock_item = SalesStock.objects.create(
                            auto_id = get_auto_id(SalesStock),
                            creator = request.user,
                            purchase_item = item.purchasestock.purchase_item,
                            qty = item.qty,
                            country = export.exporting_country,
                        )
                        stock_item.export.add(export)
                        stock_item.save()
                    
                    # stock = PurchaseStock.objects.get(purchase_item=item.purchasestock.purchase_item,is_deleted=False)
                    # stock.qty -= item.qty
                    # stock.save()
           
            response_data = {
                "status": "true",
                "title": "Successfully Updated",
                "message": "Status Update successfully.",
                'redirect': 'true',
                "redirect_url": reverse('exporting:exporting_list')
            }
                
        else:
            message =generate_form_errors(form , formset=False)
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
@login_required
@role_required(['superadmin','core_team','director'])
def print_exporting(request):
    """
    Exporting print
    :param request:
    :return: Exporting Report print
    """
    instances = Exporting.objects.filter(is_deleted=False).order_by("-date_added")
    total_qty = instances.aggregate(total_qty=Sum('exportitem__qty'))['total_qty'] or 0
    
    context = {
        'instances': instances,
        'total_qty': total_qty,
        'page_name' : 'Exporting Report',
        'page_title' : 'Exporting Report',
        'is_exporting' : True,
        'is_exporting_page': True,
    }

    return render(request, 'admin_panel/pages/export/exporting/print.html', context)

def export_exporting(request):
    filter_data = {}
    export_pk = request.GET.get("export_pk")

    exporting = Exporting.objects.filter(is_deleted=False)

    if export_pk:
        exporting = exporting.filter(pk=export_pk)

    date_range = request.GET.get('date_range')
    if date_range:
        start_date_str, end_date_str = date_range.split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        exporting = exporting.filter(date__range=[start_date, end_date])
        filter_data['date_range'] = date_range
    
    query = request.GET.get("q")
    if query:
        exporting = exporting.filter(
            Q(exporting_id__icontains=query) 
        )
        filter_data['q'] = query

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column headers
    ws.append(['#', 'Date', 'Exporting ID', 'Exporting Country', 'Courier Partner', 'Title', 'QTY', 'Total Quantity', 'Status'])

    # Iterate through Exporting objects and associated ExportItems
    for index, export in enumerate(exporting, start=1):
        courier_partner = export.courier_partner.name if export.courier_partner else ''
        exporting_country = export.exporting_country.country_name if export.exporting_country else ''

        export_id = export.exporting_id

        # Get all ExportItems for this Exporting instance
        items = export.exportitem_set.all()

        # Write the Exporting details to the worksheet
        ws.append([
            index,
            export.date,
            export_id,
            exporting_country,
            courier_partner,
            '',  # Placeholder for Title (will be filled in later)
            '',  # Placeholder for QTY (will be filled in later)
            export.total_qty(),
            export.current_status(),
        ])

        # Write each ExportItem to the worksheet
        for item in items:
            ws.append([
                '', '', '', '', '',
                item.purchasestock.purchase_item.name,
                item.qty,
                '', '',  # Placeholders for Total Quantity and Status (already filled for the main Exporting row)
            ])

    # Adjust column widths
    column_widths = [5, 15, 20, 20, 20, 20, 10, 15, 15]  # Adjust as needed
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)

    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=exporting_data.xlsx'

    return response
