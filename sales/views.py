from decimal import Decimal
import io
import json
from datetime import timezone
import datetime
from datetime import datetime
import random
#django
from django.urls import reverse
from django.db.models import Q, Sum, Min, Max
from django.db import transaction, IntegrityError
from django.contrib.auth.models import User,Group
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.forms import formset_factory, inlineformset_factory
from profit.views import calculate_profit, profit_calculation
# rest framework
from rest_framework import status
# third
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# local
from profit.models import *
from main.decorators import role_required
from sales.serializers import SalesStockSerializer
from sales.forms import DamageForm, DamageItemsForm, SalesExpenseForm, SalesForm, SalesItemsForm
from sales.models import Sales, SalesExpenses, SalesItems, SalesStock
from main.functions import generate_form_errors, get_auto_id, get_current_role, has_group

# Create your views here.
def get_sales_product_items(request):
    country_id = request.GET.get('country_id')
    product_items = SalesStock.objects.filter(qty__gt=0,country_id=country_id)
    data = [{'id': item.id, 'name': item.purchase_item.name} for item in product_items]

    return JsonResponse(data, safe=False)

def sales_item_qty(request):
    item_pk = request.GET.get("item_pk")
    country_pk = request.GET.get("country")
    
    if (instances:=SalesStock.objects.filter(pk=item_pk,country=country_pk,is_deleted=False)).exists():
        qty = instances.first().qty
        
        status_code = status.HTTP_200_OK
        response_data = {
            "status": "true",
            "qty": str(qty),
        }
    else:
        status_code = status.HTTP_404_NOT_FOUND
        response_data = {
            "status": "false",
            "title": "Failed",
            "message": "item not found",
        }

    return HttpResponse(json.dumps(response_data),status=status_code, content_type="application/json")

@login_required
@role_required(['superadmin','sales','investor'])
def sales_stock(request):
    """
    Sales Stock
    :param request:
    :return: Sales Stock List
    """
    filter_data = {}
    instances = SalesStock.objects.filter(is_deleted=False)
    
    if request.GET.get("country"):
        instances = instances.filter(country__pk=request.GET.get("country"))
        filter_data['country'] = request.GET.get("country")
        
    context = {
        'instances': instances,
        'page_name' : 'Sales Stock',
        'page_title' : 'Sales Stock',
        'filter_data': filter_data,
        
        'is_sales_pages' : True,
        'is_sales_stock_page': True,
    }

    return render(request, 'admin_panel/pages/sales/stock/list.html', context)

@login_required
@role_required(['superadmin','sales','investor'])
def sales_info(request,pk):
    """
    Sale Info
    :param request:
    :return: Sale Info single view
    """
    
    instance = Sales.objects.get(pk=pk)
    sales_items = SalesItems.objects.filter(sales=instance,is_deleted=False)
    sales_expenses = SalesExpenses.objects.filter(sales=instance,is_deleted=False)

    context = {
        'instance': instance,
        'sales_items': sales_items,
        'sales_expenses': sales_expenses,
        
        'page_name' : 'Sale Info',
        'page_title' : 'Sale Info',
    }

    return render(request, 'admin_panel/pages/sales/sales/info.html', context)

@login_required
@role_required(['superadmin','sales','investor'])
def sales_list(request):
    """
    Sales List
    :param request:
    :return: Sales List list view
    """
    filter_data = {}
    
    instances = Sales.objects.filter(is_deleted=False).order_by("-date_added")
         
    if request.GET.get('date_range'):
        start_date_str, end_date_str = request.GET.get('date_range').split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        instances = instances.filter(date__range=[start_date, end_date])
        filter_data['date_range'] = request.GET.get('date_range')
    
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(invoice_no__icontains=query) |
            Q(sales_id__icontains=query) 
        )
        title = "Sales List - %s" % query
        filter_data['q'] = query
    
    first_date_added = instances.aggregate(first_date_added=Min('date'))['first_date_added']
    last_date_added = instances.aggregate(last_date_added=Max('date'))['last_date_added']
    
    first_date_formatted = first_date_added.strftime('%m/%d/%Y') if first_date_added else None
    last_date_formatted = last_date_added.strftime('%m/%d/%Y') if last_date_added else None
    
    country = ""
    if request.GET.get("sale_country"):
        country = request.GET.get("sale_country")
        instances = instances.filter(country__pk=country)
    
    if request.GET.get("party"):
        party = request.GET.get("party")
        instances = instances.filter(sales_party__pk=party)
        filter_data['party'] = request.GET.get("party")
        
    # print(branch)
    context = {
        'instances': instances,
        'page_name' : 'Sales List',
        'page_title' : 'Sales List',
        'filter_data' :filter_data,
        'first_date_formatted': first_date_formatted,
        'last_date_formatted': last_date_formatted,
        'is_sales_pages' : True,
        'is_sales_page': True,
    }

    return render(request, 'admin_panel/pages/sales/sales/list.html', context)

@login_required
@role_required(['superadmin','Sales','director'])
def create_sales(request):
    ItemsFormset = formset_factory(SalesItemsForm, extra=2)
    ExpensesFormset = formset_factory(SalesExpenseForm, extra=2)
    
    message = ''
    if request.method == 'POST':
        form = SalesForm(request.POST)
        sales_items_formset = ItemsFormset(request.POST,prefix='sales_items_formset', form_kwargs={'empty_permitted': False})
        sales_expense_formset = ExpensesFormset(request.POST,prefix='sales_expense_formset', form_kwargs={'empty_permitted': False})
        
        if form.is_valid() and sales_items_formset.is_valid() and sales_expense_formset.is_valid():
            
            date_part = datetime.now().strftime('%Y%m%d')
            random_part = str(random.randint(1000, 9999))
            invoice_number = f'IEEIP-{date_part}-{random_part}'
            auto_id = get_auto_id(Sales)
            sales_id = "IEEIS" + str(auto_id).zfill(3)    
            
            try:
                with transaction.atomic():
                    inr_exchange_rate = ExchangeRate.objects.filter(country=form.cleaned_data['country'],is_active=True).latest('-date_added').rate_to_inr
                            
                    if form.is_valid():
                        sales_data = form.save(commit=False)
                        sales_data.auto_id = get_auto_id(Sales)
                        sales_data.creator = request.user
                        sales_data.sales_staff = request.user
                        sales_data.sales_id = sales_id
                        sales_data.invoice_no = invoice_number
                        sales_data.exchange_rate = inr_exchange_rate
                        sales_data.save()
                    
                    for form in sales_items_formset:
                        item_data = form.save(commit=False)
                        item_data.auto_id = get_auto_id(SalesItems)
                        item_data.creator = request.user
                        item_data.sales = sales_data
                        item_data.amount_in_inr = form.cleaned_data['amount'] * inr_exchange_rate
                        item_data.save()
                        
                        if item_data.sale_type == "box":
                            item_qty = item_data.qty * item_data.no_boxes
                        else:
                            item_qty = item_data.qty
                        
                        stock = SalesStock.objects.filter(country=sales_data.country,purchase_item=item_data.sales_stock.purchase_item).first()
                        if stock.qty >= Decimal(item_qty) :
                            stock.qty -= item_qty
                            stock.save()
                        else: 
                            response_data = {
                                "StatusCode": 6001,
                                "status": status.HTTP_400_BAD_REQUEST,
                                "message": "no stock available",
                            }

                        # else:
                        #     stock = SalesStock.objects.create(
                        #         auto_id = get_auto_id(SalesStock),
                        #         creator = request.user,
                        #         date_updated = datetime.datetime.today(),
                        #         updater = request.user,
                        #         country = sales_data.country,
                        #         purchase_item = item_data.sales_stock.purchase_item,
                        #         qty = item_data.qty,
                        #     )
                    
                    for form in sales_expense_formset:
                        expense_data = form.save(commit=False)
                        expense_data.auto_id = get_auto_id(SalesExpenses)
                        expense_data.creator = request.user
                        expense_data.sales = sales_data
                        expense_data.amount_in_inr = form.cleaned_data['amount'] * inr_exchange_rate
                        expense_data.save()
                    
                    # calculate_profit(sales_data.date)
                    profit_calculation(sales_data.date)
                    
                    response_data = {
                        "status": "true",
                        "title": "Successfully Created",
                        "message": "Sales created successfully.",
                        'redirect': 'true',
                        "redirect_url": reverse('sales:sales_list')
                    }
            
            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(form,formset=False)
            message += generate_form_errors(sales_items_formset,formset=True)
            message += generate_form_errors(sales_expense_formset,formset=True)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        form = SalesForm()
        sales_items_formset = ItemsFormset(prefix='sales_items_formset')
        sales_expense_formset = ExpensesFormset(prefix='sales_expense_formset')
        
        context = {
            'form': form,
            'sales_items_formset': sales_items_formset,
            'sales_expense_formset': sales_expense_formset,
            
            'page_title': 'Create Sales',
            'url': reverse('sales:create_sales'),
            'is_sales_pages' : True,
            'is_sales_page': True,
            'is_need_datetime_picker': True,
        }
        
        return render(request,'admin_panel/pages/sales/sales/create.html',context)

@login_required
@role_required(['superadmin', 'core_team', 'director'])
def edit_sales(request, pk):
    """
    Edit operation of sales
    :param request:
    :param pk:
    :return:
    """
    sales_instance = get_object_or_404(Sales, pk=pk)
    sales_items = SalesItems.objects.filter(sales=sales_instance)
    expenses = SalesExpenses.objects.filter(sales=sales_instance)

    if SalesItems.objects.filter(sales=sales_instance).exists():
        i_extra = 0
    else:
        i_extra = 1

    if SalesExpenses.objects.filter(sales=sales_instance).exists():
        e_extra = 0
    else:
        e_extra = 1

    ItemsFormset = inlineformset_factory(
        Sales,
        SalesItems,
        extra=i_extra,
        form=SalesItemsForm,
    )

    ExpensesFormset = inlineformset_factory(
        Sales,
        SalesExpenses,
        extra=e_extra,
        form=SalesExpenseForm,
    )

    message = ''

    if request.method == 'POST':
        form = SalesForm(request.POST, instance=sales_instance)
        sales_items_formset = ItemsFormset(request.POST, request.FILES,
                                           instance=sales_instance,
                                           prefix='sales_items_formset',
                                           form_kwargs={'empty_permitted': False})
        sales_expense_formset = ExpensesFormset(request.POST,
                                                instance=sales_instance,
                                                prefix='sales_expense_formset',
                                                form_kwargs={'empty_permitted': False})

        if form.is_valid() and sales_items_formset.is_valid() and sales_expense_formset.is_valid():
            try:
                with transaction.atomic():
                    
                    for sales_item in sales_items:
                        stock = SalesStock.objects.filter(country=sales_item.sales.country,purchase_item=sales_item.sales_stock.purchase_item).first()
                        if sales_item.sale_type == "box":
                            item_qty = sales_item.qty * sales_item.no_boxes
                        else:
                            item_qty = sales_item.qty
                        
                        if stock:
                            stock.qty += item_qty
                            stock.save()
                    
                    # Update sales data
                    data = form.save(commit=False)
                    data.date_updated = datetime.today()
                    data.updater = request.user
                    data.date = request.POST.get('date')
                    data.save()

                    inr_exchange_rate = ExchangeRate.objects.filter(country=data.country, is_active=True).latest(
                        '-date_added').rate_to_inr

                    # Update sales items
                    for form in sales_items_formset:
                        if form not in sales_items_formset.deleted_forms:
                            item_data = form.save(commit=False)
                            item_data.amount_in_inr = form.cleaned_data['amount'] * inr_exchange_rate
                            if not item_data.auto_id:
                                item_data.sales = sales_instance
                                item_data.auto_id = get_auto_id(SalesItems)
                                item_data.creator = request.user
                            item_data.save()
                        
                        stock = SalesStock.objects.filter(country=item_data.sales.country,purchase_item=item_data.sales_stock.purchase_item).first()
                        if sales_item.sale_type == "box":
                            item_qty = sales_item.qty * sales_item.no_boxes
                        else:
                            item_qty = sales_item.qty
                        
                        if stock:
                            stock.qty -= item_qty
                            stock.save()

                    # Delete removed sales items
                    for f in sales_items_formset.deleted_forms:
                        f.instance.delete()

                    # Update sales expenses
                    for form in sales_expense_formset:
                        if form not in sales_expense_formset.deleted_forms:
                            expense_data = form.save(commit=False)
                            if not expense_data.auto_id:
                                expense_data.auto_id = get_auto_id(SalesExpenses)
                                expense_data.creator = request.user
                                expense_data.sales = sales_instance
                            expense_data.save()

                    # Delete removed sales expenses
                    for f in sales_expense_formset.deleted_forms:
                        f.instance.delete()

                    # Calculate profit
                    profit_calculation(datetime.strptime(sales_instance.date, '%Y-%m-%d').date())

                    response_data = {
                        "status": "true",
                        "title": "Successfully Updated",
                        "message": "Sales updated successfully.",
                        'redirect': 'true',
                        "redirect_url": reverse('sales:sales_list'),
                        "return": True,
                    }

            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

        else:
            message = generate_form_errors(form, formset=False)
            message += generate_form_errors(sales_items_formset, formset=True)
            message += generate_form_errors(sales_expense_formset, formset=True)

            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')

    else:
        form = SalesForm(instance=sales_instance)
        sales_items_formset = ItemsFormset(queryset=sales_items,
                                           prefix='sales_items_formset',
                                           instance=sales_instance)
        sales_expense_formset = ExpensesFormset(queryset=expenses,
                                                prefix='sales_expense_formset',
                                                instance=sales_instance)

        context = {
            'form': form,
            'sales_items_formset': sales_items_formset,
            'sales_expense_formset': sales_expense_formset,

            'message': message,
            'page_name': 'edit sales',
            'is_sales_pages': True,
            'is_sales_page': True,
        }

        return render(request, 'admin_panel/pages/sales/sales/create.html', context)
    
@login_required
@role_required(['superadmin','sales','investor'])
def delete_sales(request, pk):
    """
    sales deletion, it only mark as is deleted field to true
    :param request:
    :param pk:
    :return:
    """
    sales = Sales.objects.get(pk=pk)
    sales_items = SalesItems.objects.filter(sales=sales)
    
    # update stock
    for item in sales_items:
        
        if item.sale_type == "box":
            item_qty = item.qty * item.no_boxes
        else:
            item_qty = item.qty
        
        stock = SalesStock.objects.get(country=sales.country,purchase_item=item.sales_stock.purchase_item)
        stock.qty += item_qty
        stock.save()
    
    SalesExpenses.objects.filter(sales=sales).update(is_deleted=True)
    
    sales_items.update(is_deleted=True)
    sales.is_deleted = True
    sales.save()
    
    profit_calculation(sales.date)
    
    response_data = {
        "status": "true",
        "title": "Successfully Deleted",
        "message": "Sale Successfully Deleted.",
        "redirect": "true",
        "redirect_url": reverse('sales:sales_list'),
    }
    
    return HttpResponse(json.dumps(response_data), content_type='application/javascript')


@login_required
@role_required(['superadmin','sales','investor'])
def print_sales(request):
    """
    Sales Print
    :param request:
    :return: Sales List Print view
    """
    filter_data = {}
    
    instances = Sales.objects.filter(is_deleted=False).order_by("-date_added")
         
    if request.GET.get('date_range'):
        start_date_str, end_date_str = request.GET.get('date_range').split(' - ')
        start_date = datetime.datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.datetime.strptime(end_date_str, '%m/%d/%Y').date()
        instances = instances.filter(date__range=[start_date, end_date])
        
        filter_data['date_range'] = request.GET.get('date_range')
    
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(invoice_no__icontains=query) |
            Q(sales_id__icontains=query) 
        )
        title = "Sales Print - %s" % query
        filter_data['q'] = query
    
        
    if request.GET.get("party"):
        party = request.GET.get("party")
        instances = instances.filter(sales_party__pk=party)
        filter_data['party'] = request.GET.get("party")
        
    # print(branch)
    context = {
        'instances': instances,
        'page_name' : 'Sales Print',
        'page_title' : 'Sales Print',
        'filter_data' :filter_data,
        'is_sales_pages' : True,
        'is_sales_page': True,
    }

    return render(request, 'admin_panel/pages/sales/sales/print.html', context)

def export_sales(request):
    filter_data = {}
    sales_pk = request.GET.get("sales_pk")

    sales = Sales.objects.filter(is_deleted=False)

    if sales_pk:
        sales = sales.filter(pk=sales_pk)

    date_range = request.GET.get('date_range')
    if date_range:
        start_date_str, end_date_str = date_range.split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        sales = sales.filter(date__range=[start_date, end_date])
        filter_data['date_range'] = date_range
    
    query = request.GET.get("q")
    if query:
        sales = sales.filter(
            Q(sales_id__icontains=query) 
        )
        filter_data['q'] = query

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column headers
    ws.append(['#', 'Date', 'Invoice No', 'Sales ID', 'Country', 'Sales Staff', 'Sales Party', 'Title', 'QTY', 'Amount', 'Amount (INR)', 'Per KG Amount', 'Total Quantity', 'Sub Total', 'Sub Total (INR)', 'Items Total Amount', 'Items Total INR Amount', 'Items Per KG Amount', 'Items Total Expense', 'Expenses Items Total INR Amount', 'Exchange Sub Total'])

    # Iterate through Sales objects
    for index, sale in enumerate(sales, start=1):
        # Get all SalesItems for this Sale instance
        items = sale.salesitems_set.all()

        # Write main sales information
        ws.append([
            index,
            sale.date,
            sale.invoice_no,
            sale.sales_id,
            sale.country.country_name if sale.country else '',
            sale.sales_staff.username,
            sale.sales_party.get_fullname() if sale.sales_party else '',
            '', '', '', '', '',  # Placeholders for item details
            sale.total_qty(),
            sale.sub_total(),
            sale.sub_total_inr(),
            sale.items_total_amount(),
            sale.items_total_inr_amount(),
            sale.items_per_kg_amount(),
            sale.items_total_expence(),
            sale.expenses_items_total_inr_amount(),
            sale.exchange_sub_total(),
        ])

        # Iterate through SalesItems and write to the worksheet
        for item_index, item in enumerate(items, start=1):
            ws.append([
                '', '', '', '', '', '', '',  # Empty placeholders for the main Sales row details
                item.sales_stock.purchase_item.name,
                item.qty,
                item.amount,
                item.amount_in_inr,
                item.per_kg_amount,
                '', '', '', '', '', '', '', '', '', '',  # Placeholders for the other sales details (already filled for the main Sales row)
            ])

    # Adjust column widths
    column_widths = [5, 15, 20, 20, 20, 20, 20,  # Adjust as needed
                     20, 10, 15, 15, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)

    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=sales_data.xlsx'

    return response

@login_required
@role_required(['superadmin','sales','investor'])
def damage_stock(request):
    """
    Damage Stock
    :param request:
    :return: Damage Stock List
    """
    filter_data = {}
    instances = DamageStock.objects.filter(is_deleted=False)
    
    if request.GET.get("country"):
        instances = instances.filter(country__pk=request.GET.get("country"))
        filter_data['country'] = request.GET.get("country")
        
    context = {
        'instances': instances,
        'page_name' : 'Damage Stock',
        'page_title' : 'Damage Stock',
        'filter_data': filter_data,
        
        'is_damage_pages' : True,
        'is_damage_stock_page': True,
    }

    return render(request, 'admin_panel/pages/sales/stock/list.html', context)

@login_required
@role_required(['superadmin','sales','investor'])
def damage_info(request,pk):
    """
    Damage Info
    :param request:
    :return: Damage Info single view
    """
    
    instance = Damage.objects.get(pk=pk)
    damage_items = DamageItems.objects.filter(damage=instance,is_deleted=False)

    context = {
        'instance': instance,
        'damage_items': damage_items,
        
        'page_name' : 'Damage Info',
        'page_title' : 'Damage Info',
    }

    return render(request, 'admin_panel/pages/sales/damage/info.html', context)

@login_required
@role_required(['superadmin','sales','investor'])
def damage_list(request):
    """
    Damage List
    :param request:
    :return: Damage List list view
    """
    filter_data = {}
    
    instances = Damage.objects.filter(is_deleted=False).order_by("-date_added")
         
    if request.GET.get('date_range'):
        start_date_str, end_date_str = request.GET.get('date_range').split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        instances = instances.filter(date__range=[start_date, end_date])
        filter_data['date_range'] = request.GET.get('date_range')
    
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(invoice_no__icontains=query) |
            Q(damage_id__icontains=query) 
        )
        title = "Damage List - %s" % query
        filter_data['q'] = query
    
    first_date_added = instances.aggregate(first_date_added=Min('date'))['first_date_added']
    last_date_added = instances.aggregate(last_date_added=Max('date'))['last_date_added']
    
    first_date_formatted = first_date_added.strftime('%m/%d/%Y') if first_date_added else None
    last_date_formatted = last_date_added.strftime('%m/%d/%Y') if last_date_added else None
    
    country = ""
    if request.GET.get("damage_country"):
        country = request.GET.get("damage_country")
        instances = instances.filter(country__pk=country)
    
    if request.GET.get("party"):
        party = request.GET.get("party")
        instances = instances.filter(damage_party__pk=party)
        filter_data['party'] = request.GET.get("party")
        
    # print(branch)
    context = {
        'instances': instances,
        'page_name' : 'Damage List',
        'page_title' : 'Damage List',
        'filter_data' :filter_data,
        'first_date_formatted': first_date_formatted,
        'last_date_formatted': last_date_formatted,
        
        'is_damage_pages' : True,
        'is_damages_page': True,
    }

    return render(request, 'admin_panel/pages/sales/damage/list.html', context)

@login_required
@role_required(['superadmin','Damage','director'])
def create_damage(request):
    ItemsFormset = formset_factory(DamageItemsForm, extra=2)
    
    message = ''
    if request.method == 'POST':
        form = DamageForm(request.POST)
        damage_items_formset = ItemsFormset(request.POST,prefix='damage_items_formset', form_kwargs={'empty_permitted': False})
        
        if form.is_valid() and damage_items_formset.is_valid():
            
            auto_id = get_auto_id(Damage)
            damage_id = "IEEIS" + str(auto_id).zfill(3)    
            
            try:
                with transaction.atomic():
                            
                    if form.is_valid():
                        damage_data = form.save(commit=False)
                        damage_data.auto_id = get_auto_id(Damage)
                        damage_data.creator = request.user
                        damage_data.damage_id = damage_id
                        damage_data.save()
                    
                    for form in damage_items_formset:
                        item_data = form.save(commit=False)
                        item_data.auto_id = get_auto_id(DamageItems)
                        item_data.creator = request.user
                        item_data.damage = damage_data
                        item_data.save()
                        
                        if (damage_stock:=DamageStock.objects.filter(country=damage_data.country,purchase_item=item_data.stock_item.purchase_item)).exists():
                            damage_stock = damage_stock.first()
                        else:
                            damage_stock = DamageStock.objects.create(
                                auto_id = get_auto_id(DamageStock),
                                creator = request.user,
                                country=damage_data.country,
                                purchase_item=item_data.stock_item.purchase_item
                                )
                            
                        if item_data.weight_type == "box":
                            item_qty = item_data.qty * item_data.no_boxes
                        else:
                            item_qty = item_data.qty
                            
                        damage_stock.qty += item_qty
                        damage_stock.save()
                        
                        sales_stock = SalesStock.objects.get(
                            country=damage_data.country,
                            purchase_item=item_data.stock_item.purchase_item
                            )
                        
                        sales_stock.qty -= item_qty
                        sales_stock.save()
                        
                    response_data = {
                        "status": "true",
                        "title": "Successfully Created",
                        "message": "Damage created successfully.",
                        'redirect': 'true',
                        "redirect_url": reverse('sales:damage_list')
                    }
            
            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }
        else:
            message = generate_form_errors(form,formset=False)
            message += generate_form_errors(damage_items_formset,formset=True)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        form = DamageForm()
        damage_items_formset = ItemsFormset(prefix='damage_items_formset')
        
        context = {
            'form': form,
            'damage_items_formset': damage_items_formset,
            
            'page_title': 'Create Damage',
            'url': reverse('sales:create_damage'),
            'is_damage_pages' : True,
            'is_damages_page': True,
            'is_need_datetime_picker': True,
        }
        
        return render(request,'admin_panel/pages/sales/damage/create.html',context)

@login_required
@role_required(['superadmin', 'sales', 'director'])
def edit_damage(request, pk):
    """
    Edit operation of damage
    :param request:
    :param pk:
    :return:
    """
    damage_instance = get_object_or_404(Damage, pk=pk)
    damage_items = DamageItems.objects.filter(damage=damage_instance)

    if DamageItems.objects.filter(damage=damage_instance).exists():
        i_extra = 0
    else:
        i_extra = 1

    ItemsFormset = inlineformset_factory(
        Damage,
        DamageItems,
        extra=i_extra,
        form=DamageItemsForm,
    )

    message = ''

    if request.method == 'POST':
        form = DamageForm(request.POST, instance=damage_instance)
        damage_items_formset = ItemsFormset(request.POST, request.FILES,
                                           instance=damage_instance,
                                           prefix='damage_items_formset',
                                           form_kwargs={'empty_permitted': False})
        
        if form.is_valid() and damage_items_formset.is_valid():
            try:
                with transaction.atomic():
                    for damage_item in damage_items:
                        if (damage_stock:=DamageStock.objects.filter(country=damage_item.damage.country,purchase_item=damage_item.stock_item.purchase_item)).exists():
                                damage_stock = damage_stock.first()
                        else:
                            damage_stock = DamageStock.objects.create(
                                auto_id = get_auto_id(DamageStock),
                                creator = request.user,
                                country=damage_item.damage.country,
                                purchase_item=damage_item.stock_item.purchase_item
                                )
                            
                        if damage_item.weight_type == "box":
                            item_qty = damage_item.qty * damage_item.no_boxes
                        else:
                            item_qty = damage_item.qty
                            
                        damage_stock.qty -= item_qty
                        damage_stock.save()
                        
                        sales_stock = SalesStock.objects.get(
                            country=damage_item.damage.country,
                            purchase_item=damage_item.stock_item.purchase_item
                            )
                        
                        sales_stock.qty += item_qty
                        sales_stock.save()
                    
                    # Update damage data
                    data = form.save(commit=False)
                    data.date_updated = datetime.today()
                    data.updater = request.user
                    data.date = request.POST.get('date')
                    data.save()

                    # Update damage items
                    for form in damage_items_formset:
                        if form not in damage_items_formset.deleted_forms:
                            item_data = form.save(commit=False)
                            if not item_data.auto_id:
                                item_data.damage = damage_instance
                                item_data.auto_id = get_auto_id(DamageItems)
                                item_data.creator = request.user
                            item_data.save()
                            
                            if (damage_stock:=DamageStock.objects.filter(country=data.country,purchase_item=item_data.stock_item.purchase_item)).exists():
                                damage_stock = damage_stock.first()
                            else:
                                damage_stock = DamageStock.objects.create(
                                    auto_id = get_auto_id(DamageStock),
                                    creator = request.user,
                                    country=data.country,
                                    purchase_item=item_data.stock_item.purchase_item
                                    )
                                
                            if item_data.weight_type == "box":
                                item_qty = item_data.qty * item_data.no_boxes
                            else:
                                item_qty = item_data.qty
                                
                            damage_stock.qty += item_qty
                            damage_stock.save()
                            
                            sales_stock = SalesStock.objects.get(
                                country=data.country,
                                purchase_item=item_data.stock_item.purchase_item
                                )
                            
                            sales_stock.qty -= item_qty
                            sales_stock.save()

                    # Delete removed damage items
                    for f in damage_items_formset.deleted_forms:
                        f.instance.delete()

                    response_data = {
                        "status": "true",
                        "title": "Successfully Updated",
                        "message": "Damage updated successfully.",
                        'redirect': 'true',
                        "redirect_url": reverse('sales:damage_list'),
                        "return": True,
                    }

            except IntegrityError as e:
                # Handle database integrity error
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

            except Exception as e:
                # Handle other exceptions
                response_data = {
                    "status": "false",
                    "title": "Failed",
                    "message": str(e),
                }

        else:
            message = generate_form_errors(form, formset=False)
            message += generate_form_errors(damage_items_formset, formset=True)

            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')

    else:
        form = DamageForm(instance=damage_instance)
        damage_items_formset = ItemsFormset(queryset=damage_items,
                                           prefix='damage_items_formset',
                                           instance=damage_instance)

        context = {
            'form': form,
            'damage_items_formset': damage_items_formset,

            'message': message,
            'page_name': 'edit damage',
            'is_damage_pages': True,
            'is_damage_page': True,
        }

        return render(request, 'admin_panel/pages/sales/damage/create.html', context)

    
@login_required
@role_required(['superadmin','sales','investor'])
def delete_damage(request, pk):
    """
    damage deletion, it only mark as is deleted field to true
    :param request:
    :param pk:
    :return:
    """
    damage = Damage.objects.get(pk=pk)
    damage_items = DamageItems.objects.filter(damage=damage)
    
    # update stock
    for item in damage_items:
        damage_stock = DamageStock.objects.get(
            country=damage.country,
            purchase_item=item.stock_item.purchase_item
            )
        
        damage_stock.qty -= item.qty
        damage_stock.save()
        
        sales_stock = SalesStock.objects.get(
            country=damage.country,
            purchase_item=item.stock_item.purchase_item
            )
        
        sales_stock.qty += item.qty
        sales_stock.save()
    
    damage_items.update(is_deleted=True)
    damage.is_deleted = True
    damage.save()
    
    response_data = {
        "status": "true",
        "title": "Successfully Deleted",
        "message": "Damage Successfully Deleted.",
        "redirect": "true",
        "redirect_url": reverse('sales:damage_list'),
    }
    
    return HttpResponse(json.dumps(response_data), content_type='application/javascript')


@login_required
@role_required(['superadmin','sales','investor'])
def print_damage(request):
    """
    Damage Print
    :param request:
    :return: Damage List Print view
    """
    filter_data = {}
    
    instances = Damage.objects.filter(is_deleted=False).order_by("-date_added")
         
    if request.GET.get('date_range'):
        start_date_str, end_date_str = request.GET.get('date_range').split(' - ')
        start_date = datetime.datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.datetime.strptime(end_date_str, '%m/%d/%Y').date()
        instances = instances.filter(date__range=[start_date, end_date])
        
        filter_data['date_range'] = request.GET.get('date_range')
    
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(invoice_no__icontains=query) |
            Q(damage_id__icontains=query) 
        )
        title = "Damage Print - %s" % query
        filter_data['q'] = query
    
        
    if request.GET.get("party"):
        party = request.GET.get("party")
        instances = instances.filter(damage_party__pk=party)
        filter_data['party'] = request.GET.get("party")
        
    # print(branch)
    context = {
        'instances': instances,
        'page_name' : 'Damage Print',
        'page_title' : 'Damage Print',
        'filter_data' :filter_data,
        'is_damage_pages' : True,
        'is_damage_page': True,
    }

    return render(request, 'admin_panel/pages/sales/damage/print.html', context)

def export_damage(request):
    damage_pk = request.GET.get("damage_pk")
    date_range = request.GET.get('date_range')
    query = request.GET.get("q")

    # Filtering Damage objects
    damage = Damage.objects.filter(is_deleted=False)
    if damage_pk:
        damage = damage.filter(pk=damage_pk)
    if date_range:
        start_date_str, end_date_str = date_range.split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        damage = damage.filter(date__range=[start_date, end_date])
    if query:
        damage = damage.filter(Q(damage_id__icontains=query))

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column headers
    ws.append(['#', 'Date', 'Damage ID', 'Country', 'Title', 'Weight Type', 'No.Boxes', 'QTY', 'Total QTY'])

    # Iterate through Damage objects
    for index, dmg in enumerate(damage, start=1):
        items = dmg.damageitems_set.all()

        # Write main damage information
        ws.append([
            index,
            dmg.date,
            dmg.damage_id,
            dmg.country.country_name if dmg.country else '',
            '', '', '', '',  # Placeholders for item details
            dmg.total_qty(),
        ])

        # Iterate through DamageItems and write to the worksheet
        for item in items:
            ws.append([
                '', '', '', '',  # Empty placeholders for the main Damage row details
                item.stock_item.purchase_item.name,
                item.get_weight_type_display(),
                item.no_boxes,
                item.qty,
                '', '', '', '',  # Placeholders for the other damage details (already filled for the main Damage row)
            ])

    # Adjust column widths
    column_widths = [5, 15, 20, 20, 20, 20, 20, 20, 10]
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Create an in-memory output file for the new workbook
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Set up the HttpResponse
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=damage_data.xlsx'
    return response