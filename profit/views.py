import io
import json
import datetime
from datetime import datetime,timedelta
#django
from django.urls import reverse
from django.db.models import Q, Sum,Min,Max
from django.http import HttpResponse
from django.db import transaction, IntegrityError
from django.contrib.auth.models import User,Group
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.forms import formset_factory, inlineformset_factory
# rest framework
from other_expences.models import OtherExpences
from rest_framework import status
# third party
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# local
from main.decorators import role_required
from main.functions import generate_form_errors, get_auto_id, get_current_role, has_group
from . models import *
from . forms import *

# Create your views here.
def calculate_profit(issued_date):
    # Filter purchases for the given date
    purchases = Purchase.objects.filter(date=issued_date, is_deleted=False)
    
    # Calculate total purchase amount
    purchase_total = PurchasedItems.objects.filter(purchase__in=purchases, is_deleted=False).aggregate(total=Sum('amount'))['total'] or 0
    
    # Filter purchase expenses for the given date
    purchase_expenses = PurchaseExpense.objects.filter(purchase__date=issued_date, is_deleted=False)
    
    # Calculate total purchase expenses amount
    purchase_expense_total = purchase_expenses.aggregate(total=Sum('amount'))['total'] or 0
    
    # Filter sales for the given date
    sales = Sales.objects.filter(date=issued_date, is_deleted=False)
    
    # Calculate total sales amount
    sales_total = sales.aggregate(total=Sum('salesitems__amount_in_inr'))['total'] or 0
    
    # Filter sales expenses for the given date
    sales_expenses = SalesExpenses.objects.filter(sales__date=issued_date, sales__is_deleted=False)
    
    # Calculate total sales expenses amount
    sales_expense_total = sales_expenses.aggregate(total=Sum('amount_in_inr'))['total'] or 0
    
    # Calculate the profit
    profit = sales_total - purchase_total - purchase_expense_total - sales_expense_total
    
    # Update or create DailyProfit instance
    instance, created = DialyProfit.objects.get_or_create(date_added=issued_date)
    instance.purchase = purchase_total
    instance.purchase_expenses = purchase_expense_total
    instance.sales = sales_total
    instance.sales_expenses = sales_expense_total
    instance.total_expenses = purchase_total + purchase_expense_total +sales_expense_total
    instance.profit = profit
    instance.save()
    # print(profit)
    # return profit

    if profit < instance.total_expenses:
        balance_profit(issued_date)

def balance_profit(issued_date):
    # Get the first day of the current month
    instance, created = DialyProfit.objects.get_or_create(date_added=issued_date)
    first_day_of_month = issued_date.replace(day=1)

    # Get the last day of the previous month
    last_day_of_previous_month = first_day_of_month - timedelta(days=1)

    # Get profits of the previous month
    previous_month_profits = DialyProfit.objects.filter(
        date_added__year=last_day_of_previous_month.year,
        date_added__month=last_day_of_previous_month.month
    ).order_by('-date_added')

    # Iterate over profits and try to balance profit
    for prev_instance in previous_month_profits:
        if prev_instance.profit > 0:  # Ensure profit is not a loss
            remaining_profit = prev_instance.profit - prev_instance.total_expenses
            shortfall = instance.total_expenses - instance.profit
            amount_to_balance = min(shortfall, remaining_profit)
            prev_instance.profit -= amount_to_balance
            instance.profit += amount_to_balance
            prev_instance.save()
            instance.save()
            break  # Stop when profit is balanced
        
def calculate_monthly_profit(year, month):
    # Get the first and last date of the month
    first_day_of_month = datetime(year, month, 1)
    next_month = first_day_of_month.replace(day=28) + timedelta(days=4)  # Ensure it's past the end of the month
    last_day_of_month = next_month - timedelta(days=next_month.day)

    # Calculate the monthly profit from DialyProfit model
    monthly_profit = DialyProfit.objects.filter(
        date_added__gte=first_day_of_month,
        date_added__lte=last_day_of_month
    ).aggregate(monthly_profit=Sum('profit'))['monthly_profit'] or 0

    # Subtract other expenses for the same month
    other_expenses = OtherExpences.objects.filter(
        date_added__year=year,
        date_added__month=month
    ).aggregate(total_expenses=Sum('amount'))['total_expenses'] or 0

    # Calculate final monthly profit
    final_monthly_profit = monthly_profit - other_expenses

    # Save the calculated monthly profit to MonthlyProfit model
    if MonthlyProfit.objects.filter(year=year, month=month).exists():
        monthly_profit_instance = MonthlyProfit.objects.get(year=year, month=month)
    else:
        monthly_profit_instance = MonthlyProfit.objects.create(
            year=year, 
            month=month,
            date_added=datetime.today().date()
            )
    
    monthly_profit_instance.other_expences = other_expenses
    monthly_profit_instance.total_revenue = monthly_profit
    monthly_profit_instance.profit = final_monthly_profit
    monthly_profit_instance.save()

def distribute_profits(year, month):
    monthly_profit = MonthlyProfit.objects.get(year=year, month=month)
    
    # core_team_users = User.objects.filter(groups__name='core_team')
    # for core_team in core_team_users:
    #     instance = CoreTeam.objects.get(user=core_team,is_deleted=False)
    #     my_profit = monthly_profit.profit * instance.share_persentage / 100
        
    #     profit_instance, created = MyProfit.objects.get_or_create(year=year, month=month, user=core_team)
    #     profit_instance.profit = my_profit
    #     profit_instance.save()
        
    investor_users = User.objects.filter(groups__name='investor').exclude(username__endswith='_deleted')
    for investor in investor_users:
        instance = Investors.objects.get(user=investor,is_deleted=False)
        my_profit = monthly_profit.profit * instance.share_persentage / 100
        
        if MyProfit.objects.filter(year=year, month=month, user=investor).exists():
            profit_instance = MyProfit.objects.get(year=year, month=month, user=investor)
        else:
            profit_instance = MyProfit.objects.create(date_added=datetime.today().date(),year=year, month=month, user=investor)
        
        profit_instance.profit = my_profit
        profit_instance.save()
        
@login_required
@role_required(['superadmin','core_team','director'])
def exchange_rates(request):
    """
    Exchange Rate
    :param request:
    :return: Exchange Rate List
    """
    instances = ExchangeRate.objects.filter(is_deleted=False)
    
    context = {
        'instances': instances,
        'page_name' : 'Exchange Rate',
        'page_title' : 'Exchange Rate',
        'is_exchange_rate' : True,
        'is_exchange_rates_page': True,
    }

    return render(request, 'admin_panel/pages/exchange/list.html', context)

@login_required
@role_required(['superadmin','core_team','director'])
def create_exchange_rate(request):
    
    message = ''
    if request.method == 'POST':
        form = ExchangeRateForm(request.POST)
        now = timezone.now()
        
        if form.is_valid() :
            data = form.save(commit=False)
            data.auto_id = get_auto_id(ExchangeRate)
            data.creator = request.user
            if data.start_date <= now <= data.end_date:
                data.is_active = True
            else:
                data.is_active = False
            data.save()
                    
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Exchange Rate created successfully.",
                'redirect': 'true',
                "redirect_url": reverse('profit:exchange_rates')
            }
    
        else:
            message = generate_form_errors(form, formset=False)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        form = ExchangeRateForm()
        
        context = {
            'form': form,
            
            'page_title': 'Create Exchange Rate',
            'url': reverse('profit:create_exchange_rate'),
            'is_exchange_rate' : True,
            'is_exchange_rates_page': True,
        }
        
        return render(request,'admin_panel/pages/exchange/create.html',context)
    
@login_required
@role_required(['superadmin','core_team','director'])
def edit_exchange_rate(request,pk):
    """
    edit operation of export item
    :param request:
    :param pk:
    :return:
    """
    instance = get_object_or_404(ExchangeRate, pk=pk)
        
    message = ''
    if request.method == 'POST':
        form = ExchangeRateForm(request.POST,instance=instance)
        now = timezone.now()
        
        if form.is_valid():
            
            data = form.save(commit=False)
            data.date_updated = datetime.today()
            data.updater = request.user
            if data.start_date <= now <= data.end_date:
                data.is_active = True
            else:
                data.is_active = False
            data.save()
                    
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Purchase Item Update successfully.",
                'redirect': 'true',
                "redirect_url": reverse('profit:exchange_rates')
            }
    
        else:
            message = generate_form_errors(form ,formset=False)
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        
        form = ExchangeRateForm(instance=instance)

        context = {
            'form': form,
            'page_name' : 'Edit Exchange Rate',
            'page_title' : 'Edit Exchange Rate',
            'url' : reverse('profit:create_exchange_rate'),
            'is_exchange_rate' : True,
            'is_exchange_rates_page': True,  
        }

        return render(request, 'admin_panel/pages/exchange/create.html',context)

@login_required
@role_required(['superadmin','core_team','director'])
def delete_exchange_rate(request, pk):
    """
    exchange rates deletion, it only mark as is deleted field to true
    :param request:
    :param pk:
    :return:
    """
    invoice_numbers_text = ""
    # data = ExchangeRate.objects.get(pk=pk)
    # if not SalesMaterials.objects.filter(export_item=data, is_deleted=False).exists():
        
    #     data.is_deleted=True
    #     data.save()
    
    #     if (instances:=PurchaseMaterials.objects.filter(export_item=data)).exists():
    #         instances.update(is_deleted=True)
            
    #     if (stocks:=Stock.objects.filter(export_item=data)).exists():
    #         stocks.update(is_deleted=True)

    #     response_data = {
    #         "status": "true",
    #         "title": "Successfully Deleted",
    #         "message": "Purchase Item Successfully Deleted.",
    #         "redirect": "true",
    #         "redirect_url": reverse('profit:exchange_rates'),
    #     }
    # else:
    #     sales_material_queryset = SalesMaterials.objects.filter(export_item=data)
    #     sales_invoice_numbers = [sales_material.sales.invoice_no for sales_material in sales_material_queryset]
    #     unique_invoice_numbers = list(set(sales_invoice_numbers))
    #     invoice_numbers_text = ", ".join(unique_invoice_numbers)

    message = f"This export item has already included the sale of some items. Sales Invoice Numbers: {invoice_numbers_text}"
    
    response_data = {
        "status": "false",
        "title": "Failed",
        "message": message,
    }

    return HttpResponse(json.dumps(response_data), content_type='application/javascript')

@login_required
@role_required(['superadmin','core_team','director'])
def print_exchange_rates(request):
    """
    Print Exchange Rate
    :param request:
    :return: Print Exchange Rate List
    """
    instances = ExchangeRate.objects.filter(is_deleted=False)
    
    context = {
        'instances': instances,
        'page_name' : 'Exchange Rate',
        'page_title' : 'Exchange Rate',
        'is_exchange_rate' : True,
        'is_exchange_rates_page': True,
    }

    return render(request, 'admin_panel/pages/exchange/print.html', context)

def export_exchange_rates(request):
    # Fetch all Exchange Rates
    exchange_rates = ExchangeRate.objects.all()

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column headers
    ws.append(['#', 'Country', 'Currency', 'Rate in INR','Start Date','End Date','Is Active'])

    # Iterate through Exchange Rates and write to the worksheet
    for index, rate in enumerate(exchange_rates, start=1):
        ws.append([
            index,
            rate.country.country_name if rate.country else '',
            rate.currency,
            float(rate.rate_to_inr),  # Convert DecimalField to float
            rate.start_date.strftime("%Y-%m-%d %H:%M:%S"),  # Convert DateTimeField to string
            rate.end_date.strftime("%Y-%m-%d %H:%M:%S"),    # Convert DateTimeField to string
            "Yes" if rate.is_active else "No",  # Convert BooleanField to string
        ])

    # Adjust column widths
    column_widths = [5, 30, 15, 20, 20, 20, 10]  # Adjust as needed
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Prepare the response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=exchange_rates.xlsx'

    return response


@login_required
@role_required(['superadmin','core_team','director'])
def dialy_profits(request):
    """
    Profits
    :param request:
    :return: Profit List
    """
    filter_data = {}
    
    instances = DialyProfit.objects.all().order_by("-date_added")
    
    date_range = request.GET.get('date_range')

    if date_range:
        start_date_str, end_date_str = date_range.split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        instances = instances.filter(date_added__range=[start_date, end_date])
        filter_data['date_range'] = date_range
    
    first_date_added = instances.aggregate(first_date_added=Min('date_added'))['first_date_added']
    last_date_added = instances.aggregate(last_date_added=Max('date_added'))['last_date_added']
    
    first_date_formatted = first_date_added.strftime('%m/%d/%Y') if first_date_added else None
    last_date_formatted = last_date_added.strftime('%m/%d/%Y') if last_date_added else None
    
    context = {
        'instances': instances,
        'first_date_formatted': first_date_formatted,
        'last_date_formatted': last_date_formatted,
        'filter_data' :filter_data,
        
        'page_name' : 'Dialy Profit List',
        'page_title' : 'Dialy Profit List',
        'is_profit' : True,
        'is_dialy_profit_page': True,
    }

    return render(request, 'admin_panel/pages/profit/dialy_list.html', context)

@login_required
@role_required(['superadmin','core_team','director'])
def print_dialy_profits(request):
    """
    Profits print
    :param request:
    :return: print Profit List
    """
    instances = DialyProfit.objects.all().order_by("-date_added")
    
    context = {
        'instances': instances,
        
        'page_name' : 'Dialy Profit List',
        'page_title' : 'Dialy Profit List',
        'is_profit' : True,
        'is_dialy_profit_page': True,
    }

    return render(request, 'admin_panel/pages/profit/print_dialy_list.html', context)

@login_required
@role_required(['superadmin','core_team','director','investor'])
def export_dialy_profits(request):
    filter_data = {}
    profit_pk = request.GET.get("profit_pk")

    dialy_profits = DialyProfit.objects.all()

    if profit_pk:
        dialy_profits = dialy_profits.filter(pk=profit_pk)
        
    date_range = request.GET.get('date_range')
    if date_range:
        start_date_str, end_date_str = date_range.split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        dialy_profits = dialy_profits.filter(date_added__range=[start_date, end_date])
        filter_data['date_range'] = date_range
    
    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column headers
    ws.append(['Date','purchase','purchase_expenses','sales','sales_expenses','total_expenses','profit'])

    # Fetch and write data to Excel
    for profit in dialy_profits:
        ws.append([
            profit.date_added.date(),
            profit.purchase,
            profit.purchase_expenses,
            profit.sales,
            profit.sales_expenses,
            profit.total_expenses,
            profit.profit
        ])

    # Prepare response
    output = io.BytesIO()
    wb.save(output)

    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=dialy_profits_data.xlsx'

    return response


@login_required
@role_required(['superadmin','core_team','director'])
def monthly_profits(request):
    """
    Profits
    :param request:
    :return: Profit List
    """
    filter_data = {}
    
    instances = MonthlyProfit.objects.all().order_by("-date_added")
    
    date_range = request.GET.get('date_range')

    if date_range:
        start_date_str, end_date_str = date_range.split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        instances = instances.filter(date_added__range=[start_date, end_date])
        filter_data['date_range'] = date_range
    
    first_date_added = instances.aggregate(first_date_added=Min('date_added'))['first_date_added']
    last_date_added = instances.aggregate(last_date_added=Max('date_added'))['last_date_added']
    
    first_date_formatted = first_date_added.strftime('%m/%d/%Y') if first_date_added else None
    last_date_formatted = last_date_added.strftime('%m/%d/%Y') if last_date_added else None
    
    context = {
        'instances': instances,
        'first_date_formatted': first_date_formatted,
        'last_date_formatted': last_date_formatted,
        'filter_data' :filter_data,
        
        'page_name' : 'Monthly Profit List',
        'page_title' : 'Monthly Profit List',
        'is_profit' : True,
        'is_monthly_profit_page': True,
    }

    return render(request, 'admin_panel/pages/profit/monthly_list.html', context)

@login_required
@role_required(['superadmin','core_team','director'])
def print_monthly_profits(request):
    """
    Profits
    :param request:
    :return: Print Profit List
    """
    instances = MonthlyProfit.objects.all().order_by("-date_added")
    
    
    context = {
        'instances': instances,
        
        'page_name' : 'Monthly Profit List',
        'page_title' : 'Monthly Profit List',
        'is_profit' : True,
        'is_monthly_profit_page': True,
    }

    return render(request, 'admin_panel/pages/profit/print_monthly_list.html', context)

@login_required
@role_required(['superadmin','core_team','director','investor'])
def export_monthly_profits(request):
    filter_data = {}
    profit_pk = request.GET.get("profit_pk")

    monthly_profits = MonthlyProfit.objects.all()

    if profit_pk:
        monthly_profits = monthly_profits.filter(pk=profit_pk)
        
    date_range = request.GET.get('date_range')
    if date_range:
        start_date_str, end_date_str = date_range.split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        monthly_profits = monthly_profits.filter(date_added__range=[start_date, end_date])
        filter_data['date_range'] = date_range
    
    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column headers
    ws.append(['Date Added','Year','Month','Total Revenue','Other Expences','Profit'])

    # Fetch and write data to Excel
    for profit in monthly_profits:
        ws.append([
            profit.date_added.date(),
            profit.year,
            profit.month,
            profit.total_revenue,
            profit.other_expences,
            profit.profit
        ])

    # Prepare response
    output = io.BytesIO()
    wb.save(output)

    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=monthly_profits_data.xlsx'

    return response

@login_required
@role_required(['superadmin','core_team','director','investor'])
def users_profits(request):
    """
    Profits
    :param request:
    :return: Profit List
    """
    filter_data = {}
    
    instances = MyProfit.objects.all()
    
    if request.user.groups.filter(name__in=['core_team','investor']).exists():
        instances = instances.filter(user=request.user)
        
    date_range = request.GET.get('date_range')

    if date_range:
        start_date_str, end_date_str = date_range.split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        instances = instances.filter(date_added__range=[start_date, end_date])
        filter_data['date_range'] = date_range
        
    if request.GET.get('investor_pk'):
        instances = instances.filter(user__pk=request.GET.get('investor_pk'))
        filter_data['investor_pk'] = request.GET.get('investor_pk')
    
    first_date_added = MyProfit.objects.aggregate(first_date_added=Min('date_added'))['first_date_added']
    last_date_added = MyProfit.objects.aggregate(last_date_added=Max('date_added'))['last_date_added']
    
    first_date_formatted = first_date_added.strftime('%m/%d/%Y') if first_date_added else None
    last_date_formatted = last_date_added.strftime('%m/%d/%Y') if last_date_added else None
    
    context = {
        'instances': instances.order_by("-date_added"),
        'first_date_formatted': first_date_formatted,
        'last_date_formatted': last_date_formatted,
        'filter_data' :filter_data,
        
        'page_name' : 'My Profit List',
        'page_title' : 'My Profit List',
        'is_profit' : True,
        'is_my_profit_page': True,
    }

    return render(request, 'admin_panel/pages/profit/my_list.html', context)

@login_required
@role_required(['superadmin','core_team','director'])
def print_my_profits(request):
    """
    Profits
    :param request:
    :return: Print Profit List
    """
    instances = MyProfit.objects.all().order_by("-date_added")
    
    context = {
        'instances': instances,
        
        'page_name' : 'Investers Profit List',
        'page_title' : 'Investers Profit List',
        'is_profit' : True,
        'is_my_profit_page': True,
    }

    return render(request, 'admin_panel/pages/profit/print_my_list.html', context)

@login_required
@role_required(['superadmin','core_team','director','investor'])
def export_my_profits(request):
    filter_data = {}
    profit_pk = request.GET.get("profit_pk")

    investors_profits = MyProfit.objects.all()

    if profit_pk:
        investors_profits = investors_profits.filter(pk=profit_pk)
        
    date_range = request.GET.get('date_range')
    if date_range:
        start_date_str, end_date_str = date_range.split(' - ')
        start_date = datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.strptime(end_date_str, '%m/%d/%Y').date()
        investors_profits = investors_profits.filter(date_added__range=[start_date, end_date])
        filter_data['date_range'] = date_range
    
    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column headers
    ws.append(['Date Added','Year','Month','Full Name','Profit'])

    # Fetch and write data to Excel
    for profit in investors_profits:
        ws.append([
            profit.date_added.date(),
            profit.year,
            profit.month,
            profit.get_username(),
            profit.profit
        ])

    # Prepare response
    output = io.BytesIO()
    wb.save(output)

    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=investors_profits_data.xlsx'

    return response