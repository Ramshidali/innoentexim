#standerd
import io
import json
import datetime
#django
from django.db.models import Q, Max, Min
from django.urls import reverse
from django.conf import settings
from django.db.models import Sum
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
# third party
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
#local
from . forms import *
from . models import *
from main.decorators import role_required
from main.functions import generate_form_errors, get_auto_id

# Create your views here.
@login_required
@role_required(['superadmin','core_team','director'])
def other_expence_type_list(request):
    """
    ExpenceTypes listings
    :param request:
    :return: ExpenceTypes list view
    """
    instances = ExpenceTypes.objects.filter(is_deleted=False).order_by("-id")
    
    filter_data = {}
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(auto_id__icontains=query) |
            Q(name__icontains=query) 
        )
        title = "ExpenceTypes - %s" % query
        filter_data['q'] = query
    

    context = {
        'instances': instances,
        'page_name' : 'Expence Types',
        'page_title' : 'Expence Types',
        'filter_data' :filter_data,
        'is_other_expences' : True,
        'is_other_expence_type_page' : True, 
    }

    return render(request, 'admin_panel/pages/other_expences/types/list.html', context)

@login_required
@role_required(['superadmin','core_team','director'])
def create_other_expence_type(request):
    """
    create operation of ExpenceTypes
    :param request:
    :param pk:
    :return:
    """
    if request.method == 'POST':
        # if instance go to edit
        form = ExpenceTypesForm(request.POST)
            
        if form.is_valid():
            data = form.save(commit=False)
            data.auto_id = get_auto_id(ExpenceTypes)
            data.creator = request.user
            data.date_updated = datetime.datetime.today()
            data.updater = request.user
            data.save()

            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Expence Types created successfully.",
                'redirect': 'true',
                "redirect_url": reverse('other_expences:other_expence_type_list')
            }
    
        else:
            message =generate_form_errors(form , formset=False)
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        
        form = ExpenceTypesForm()

        context = {
            'form': form,
            'page_name' : 'Create Expence Types',
            'page_title' : 'Create Expence Types',
            'is_other_expences' : True,
            'is_other_expence_type_page' : True, 
        }

        return render(request, 'admin_panel/pages/other_expences/types/create.html',context)
    
@login_required
@role_required(['superadmin','core_team','director'])
def edit_other_expence_type(request,pk):
    """
    edit operation of other_expences
    :param request:
    :param pk:
    :return:
    """
    instance = get_object_or_404(ExpenceTypes, pk=pk)
        
    message = ''
    if request.method == 'POST':
        form = ExpenceTypesForm(request.POST,files=request.FILES,instance=instance)
        
        if form.is_valid():
            #update ExpenceTypes
            data = form.save(commit=False)
            data.date_updated = datetime.datetime.today()
            data.updater = request.user
            data.save()
                    
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Expence Types Update successfully.",
                'redirect': 'true',
                "redirect_url": reverse('other_expences:other_expence_type_list'),
            }
    
        else:
            message = generate_form_errors(form ,formset=False)
            
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        
        form = ExpenceTypesForm(instance=instance)

        context = {
            'form': form,
            'page_name' : 'Create Expence Types',
            'page_title' : 'Create Expence Types',
            'is_other_expences' : True,
            'is_other_expence_type_page' : True,
        }

        return render(request, 'admin_panel/pages/other_expences/types/create.html',context)

@login_required
@role_required(['superadmin','core_team','director'])
def delete_other_expence_type(request, pk):
    """
    ExpenceTypes deletion, it only mark as is deleted field to true
    :param request:
    :param pk:
    :return:
    """
    ExpenceTypes.objects.filter(pk=pk).update(is_deleted=True)

    response_data = {
        "status": "true",
        "title": "Successfully Deleted",
        "message": "Expence Types Successfully Deleted.",
        "redirect": "true",
        "redirect_url": reverse('other_expences:other_expence_type_list'),
    }

    return HttpResponse(json.dumps(response_data), content_type='application/javascript')


@login_required
@role_required(['superadmin','core_team','director'])
def other_expence_list(request):
    """
    OtherExpences listings
    :param request:
    :return: OtherExpences list view
    """
    instances = OtherExpences.objects.filter(is_deleted=False).order_by("-date_added")
    
    
    filter_data = {}
    query = request.GET.get("q")
    
    date_range = ""
    if request.GET.get('date_range'):
        start_date_str, end_date_str = request.GET.get('date_range').split(' - ')
        start_date = datetime.datetime.strptime(start_date_str, '%m/%d/%Y').date()
        end_date = datetime.datetime.strptime(end_date_str, '%m/%d/%Y').date()
        instances = instances.filter(date_added__range=[start_date, end_date])
        filter_data['date_range'] = request.GET.get('date_range')
    
    if query:

        instances = instances.filter(
            Q(expence_type__name__icontains=query) 
        )
        title = "Other Expences Report - %s" % query
        filter_data['q'] = query
        
    first_date_added = instances.aggregate(first_date_added=Min('date_added'))['first_date_added']
    last_date_added = instances.aggregate(last_date_added=Max('date_added'))['last_date_added']
    
    first_date_formatted = first_date_added.strftime('%m/%d/%Y') if first_date_added else None
    last_date_formatted = last_date_added.strftime('%m/%d/%Y') if last_date_added else None
    
    if request.GET.get("expence_type"):
        instances = instances.filter(expence_type__pk=request.GET.get("expence_type"))
        filter_data['expence_type'] = request.GET.get("expence_type")

    total_amount = instances.aggregate(total_amount=Sum('amount'))['total_amount'] or 0
    context = {
        'instances': instances,
        'total_amount': total_amount,
        'page_name' : 'Other Expence',
        'page_title' : 'Other Expence',
        'filter_data' :filter_data,
        'first_date_formatted': first_date_formatted,
        'last_date_formatted': last_date_formatted,
        'is_other_expences' : True,
        'is_other_expence_page' : True, 
    }

    return render(request, 'admin_panel/pages/other_expences/list.html', context)

@login_required
@role_required(['superadmin','core_team','director'])
def create_other_expence(request):
    """
    create operation of Other Expence
    :param request:
    :param pk:
    :return:
    """
    if request.method == 'POST':
        # if instance go to edit
        form = OtherExpencesForm(request.POST)
            
        if form.is_valid():
            data = form.save(commit=False)
            data.auto_id = get_auto_id(OtherExpences)
            data.creator = request.user
            data.date_updated = datetime.datetime.today()
            data.updater = request.user
            data.save()

            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Expence Types created successfully.",
                'redirect': 'true',
                "redirect_url": reverse('other_expences:other_expence_list')
            }
    
        else:
            message =generate_form_errors(form , formset=False)
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message,
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        
        form = OtherExpencesForm()

        context = {
            'form': form,
            'page_name' : 'Create Expence Types',
            'page_title' : 'Create Expence Types',
            'is_other_expences' : True,
            'is_other_expence_page' : True, 
        }

        return render(request, 'admin_panel/pages/other_expences/create.html',context)
    
@login_required
@role_required(['superadmin','core_team','director'])
def edit_other_expence(request,pk):
    """
    edit operation of other_expences
    :param request:
    :param pk:
    :return:
    """
    instance = get_object_or_404(OtherExpences, pk=pk)
        
    message = ''
    if request.method == 'POST':
        form = OtherExpencesForm(request.POST,files=request.FILES,instance=instance)
        
        if form.is_valid():
            #update OtherExpences
            data = form.save(commit=False)
            data.date_updated = datetime.datetime.today()
            data.updater = request.user
            data.save()
                    
            response_data = {
                "status": "true",
                "title": "Successfully Created",
                "message": "Expence Types Update successfully.",
                'redirect': 'true',
                "redirect_url": reverse('other_expences:other_expence_list'),
            }
    
        else:
            message = generate_form_errors(form ,formset=False)
            
            
            response_data = {
                "status": "false",
                "title": "Failed",
                "message": message
            }

        return HttpResponse(json.dumps(response_data), content_type='application/javascript')
    
    else:
        
        form = OtherExpencesForm(instance=instance)

        context = {
            'form': form,
            'page_name' : 'Create Expence Types',
            'page_title' : 'Create Expence Types',
            'is_other_expences' : True,
            'is_other_expence_page' : True,
        }

        return render(request, 'admin_panel/pages/other_expences/create.html',context)

@login_required
@role_required(['superadmin','core_team','director'])
def delete_other_expence(request, pk):
    """
    OtherExpences deletion, it only mark as is deleted field to true
    :param request:
    :param pk:
    :return:
    """
    OtherExpences.objects.filter(pk=pk).update(is_deleted=True)

    response_data = {
        "status": "true",
        "title": "Successfully Deleted",
        "message": "Expence Successfully Deleted.",
        "redirect": "true",
        "redirect_url": reverse('other_expences:other_expence_list'),
    }

    return HttpResponse(json.dumps(response_data), content_type='application/javascript')

@login_required
@role_required(['superadmin','core_team','director'])
def print_other_expenses(request):
    """
    Other Expences Print
    :param request:
    :return: Other Expences print view
    """
    instances = OtherExpences.objects.filter(is_deleted=False).order_by("-date_added")
    
    total_amount = instances.aggregate(total_amount=Sum('amount'))['total_amount'] or 0
    
    filter_data = {}
    query = request.GET.get("q")
    
    if query:

        instances = instances.filter(
            Q(auto_id__icontains=query) |
            Q(name__icontains=query) 
        )
        title = "ExpenceTypes - %s" % query
        filter_data['q'] = query
    

    context = {
        'instances': instances,
        'total_amount': total_amount,
        'page_name' : 'Other Expence',
        'page_title' : 'Other Expence',
        'filter_data' :filter_data,
        'is_other_expences' : True,
        'is_other_expence_page' : True, 
    }

    return render(request, 'admin_panel/pages/other_expences/print.html', context)

def export_other_expenses(request):
    # Fetch all other expenses
    other_expenses = OtherExpences.objects.all()

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active

    # Define column headers
    ws.append(['#', 'Remark', 'Amount', 'Expense Type'])

    # Iterate through other expenses and write to the worksheet
    for index, expense in enumerate(other_expenses, start=1):
        ws.append([
            index,
            expense.remark if expense.remark else '',
            expense.amount,
            expense.expence_type.name if expense.expence_type else '',
        ])

    # Adjust column widths
    column_widths = [5, 30, 15, 20]  # Adjust as needed
    for i, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    # Prepare the response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=other_expenses.xlsx'

    return response